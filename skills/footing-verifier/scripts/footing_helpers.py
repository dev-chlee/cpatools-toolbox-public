"""
footing_helpers.py - 풋팅 검증 보조 헬퍼

이 모듈은 **결정론적 보조 작업**만 담당한다:
- 검증 셀 서식 적용 (노란색 배경, 테두리 등)
- 서술 오류 셀에 분홍색 + 셀메모
- LibreOffice 재계산 + 오류 처리
- 검증결과요약 시트 생성 (하이퍼링크 포함)
- 시트 참조 문자열/하이퍼링크 생성 (따옴표 처리)

위계 분석, 합계 판단, 부호 처리, 서술 적절성 같은 **회계 판단**은 
LLM이 시트를 직접 읽고 결정하며, 이 모듈을 호출하지 않는다.

사용 예:
    from footing_helpers import (
        add_check, add_header, mark_narrative_error,
        recalc_and_verify, build_summary_sheet,
        analyze_workbook_basic, sheet_ref, hyperlink,
    )
    
    # LLM이 직접 시트를 읽고 판단한 후 검증 수식을 작성
    log = []
    add_check(ws_bs, log, "BS 자산총계 당기", "G68", "=D68-(D11+D32)", "당기")
    add_check(ws_bs, log, "BS 대차균형 당기", "G98", "=D68-D97", "당기")
    # ...
    
    # 헬퍼로 서식 마무리
    recalc_and_verify("output.xlsx")
    build_summary_sheet("output.xlsx", log, narrative_findings, structure)
"""

import os
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

# ─────────────────────── 서식 상수 ───────────────────────

YELLOW = PatternFill('solid', start_color='FFFF00', end_color='FFFF00')
RED = PatternFill('solid', start_color='FF0000', end_color='FF0000')
LIGHT_RED = PatternFill('solid', start_color='FFCCCC', end_color='FFCCCC')
LIGHT_GREEN = PatternFill('solid', start_color='C6EFCE', end_color='C6EFCE')
LIGHT_AMBER = PatternFill('solid', start_color='FFEB9C', end_color='FFEB9C')
HEADER_FILL = PatternFill('solid', start_color='305496', end_color='305496')
SUB_HEADER_FILL = PatternFill('solid', start_color='8EA9DB', end_color='8EA9DB')
COL_HEADER_FILL = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')

WHITE_BOLD = Font(color='FFFFFF', bold=True, name='맑은 고딕', size=11)
WHITE_BOLD_S = Font(color='FFFFFF', bold=True, name='맑은 고딕', size=10)
NORMAL = Font(name='맑은 고딕', size=10)
BOLD = Font(name='맑은 고딕', size=10, bold=True)
LINK = Font(name='맑은 고딕', size=10, color='0563C1', underline='single')
ERR_FONT = Font(name='맑은 고딕', size=10, bold=True, color='C00000')
MONO = Font(name='Consolas', size=9)

THIN = Side(border_style='thin', color='808080')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

NUM_FMT = '#,##0;[Red](#,##0);-'

ERROR_VALUES = {'#VALUE!', '#REF!', '#DIV/0!', '#NAME?', '#NUM!', '#N/A', '#NULL!'}

SOFFICE_CANDIDATES = (
    'soffice',
    'libreoffice',
    r'C:\Program Files\LibreOffice\program\soffice.exe',
    r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
    '/Applications/LibreOffice.app/Contents/MacOS/soffice',
    '/usr/bin/soffice',
    '/usr/bin/libreoffice',
    '/usr/local/bin/soffice',
)


# ─────────────────────── 시트 참조 문자열 ───────────────────────

def needs_quote(sheet_name):
    """시트명이 작은따옴표를 필요로 하는지 (숫자, 공백, 특수문자 포함)"""
    if sheet_name.isdigit():
        return True
    return any(c in sheet_name for c in ' -.()&,!@#$%^*+=[]{}|;:<>?/\\\'\"')


def sheet_ref(sheet_name, cell):
    """엑셀 수식에서 시트 참조 문자열 생성: BS!D11 또는 '9'!C46"""
    if needs_quote(sheet_name):
        return f"'{sheet_name}'!{cell}"
    return f"{sheet_name}!{cell}"


def hyperlink(sheet_name, cell_ref, display_text):
    """=HYPERLINK 수식 생성 (시트 내 이동)"""
    if needs_quote(sheet_name):
        target = f"#'{sheet_name}'!{cell_ref}"
    else:
        target = f"#{sheet_name}!{cell_ref}"
    return f'=HYPERLINK("{target}","{display_text}")'


# ─────────────────────── 워크북 기본 정보 추출 ───────────────────────

def analyze_workbook_basic(wb):
    """
    워크북의 기본 메타정보만 추출 (시트 인벤토리, 회사명, 기수, 단위 등).
    위계 분석/소계 위치 판단은 하지 않는다 - LLM이 직접 시트를 보고 판단.
    
    Returns:
        {
            'sheets': [{'name': str, 'rows': int, 'cols': int}, ...],
            'company': '주식회사 삼성전자' 또는 None,
            'period_current': {'fiscal_year': N, 'closing_year': YYYY, 'closing_date': 'YYYY-MM-DD'} 또는 None,
            'period_prior': {...} 또는 None,
            'unit': '원' / '천원' / '백만원' / None,
            'note_sheets': [숫자 시트명 또는 '주석N' 패턴 시트명 리스트],
            'note_titles': {'4': '유형자산', '7': '지분법적용투자주식', ...},
        }
    """
    info = {
        'sheets': [],
        'company': None,
        'period_current': None,
        'period_prior': None,
        'unit': None,
        'note_sheets': [],
        'note_titles': {},
    }
    
    # 시트 인벤토리
    for name in wb.sheetnames:
        ws = wb[name]
        info['sheets'].append({
            'name': name,
            'rows': ws.max_row,
            'cols': ws.max_column,
        })
    
    # 주석 시트 식별 (숫자 또는 '주석N' 패턴)
    note_pattern = re.compile(r'^(?:주석\s*)?(\d+)$')
    for name in wb.sheetnames:
        m = note_pattern.match(name.strip())
        if m:
            note_num = m.group(1)
            info['note_sheets'].append(name)
            # 본문에서 'N. 제목' 추출
            ws = wb[name]
            for r in range(1, min(15, ws.max_row + 1)):
                for c_idx in [2, 1]:
                    v = ws.cell(row=r, column=c_idx).value
                    if not isinstance(v, str):
                        continue
                    title_m = re.match(r'^\s*' + re.escape(note_num) + r'\.\s*(.+)$', v.strip())
                    if title_m:
                        info['note_titles'][note_num] = title_m.group(1).strip()
                        break
                if note_num in info['note_titles']:
                    break
    
    # 회사명/기수/단위는 BS 또는 첫 번째 재무제표에서 추출
    fs_candidates = []
    for name in wb.sheetnames:
        if name.upper() in ['BS', 'B/S', 'PL', 'P/L', 'CE', 'CF', 'IS']:
            fs_candidates.append(name)
        elif any(kw in name for kw in ['재무상태표', '손익계산서', '자본변동표', '현금흐름표']):
            fs_candidates.append(name)
    
    if fs_candidates:
        ws = wb[fs_candidates[0]]
        info['company'] = _extract_company(ws)
        info['period_current'], info['period_prior'] = _extract_periods(ws)
        info['unit'] = _extract_unit(ws)

    return info


# ─────────────────────── 메타 추출 헬퍼 ───────────────────────

# 회사명: 한국 법인격 표기 변종 망라
_COMPANY_KEYWORDS = (
    '주식회사', '(주)', '㈜', '유한회사', '(유)',
    'Inc', 'Inc.', 'Co.', 'Co.,', 'Ltd', 'Ltd.',
    'Corp', 'Corp.', 'Corporation', 'Limited', 'Company', 'LLC',
)
_COMPANY_BLOCKLIST = ('주식회사명', '회사명',)  # 라벨 셀 (예: "회사명: ___")


def _extract_company(ws):
    """첫 15행에서 회사명 추출. 키워드 포함 셀 우선, 없으면 단독 짧은 텍스트 셀."""
    candidates = []
    for r in range(1, min(16, ws.max_row + 1)):
        for c_idx in (2, 1, 3, 4):
            v = ws.cell(row=r, column=c_idx).value
            if not isinstance(v, str):
                continue
            v_strip = v.strip()
            if not v_strip or len(v_strip) > 100:
                continue
            # 라벨 셀 제외
            if any(bl in v_strip for bl in _COMPANY_BLOCKLIST) and ':' not in v_strip:
                continue
            # 1순위: 법인격 키워드 포함
            if any(kw in v_strip for kw in _COMPANY_KEYWORDS):
                # "회사명: (주)ABC" 같은 라벨 형식이면 ':' 뒤만
                if ':' in v_strip and any(lab in v_strip.split(':')[0] for lab in ('회사명', 'Company', '상호')):
                    after = v_strip.split(':', 1)[1].strip()
                    if after:
                        return after
                return v_strip
            # 2순위 후보: 짧은 단독 텍스트 (재무제표 셀 가능성 낮음)
            if (2 <= len(v_strip) <= 30
                and not any(ch.isdigit() for ch in v_strip)
                and not any(kw in v_strip for kw in (
                    '재무상태표', '손익계산서', '자본변동표', '현금흐름표',
                    '단위', '제 ', '제(', '주석', '당기', '전기',
                    '과목', '자산', '부채', '자본', 'BS', 'PL', 'CF', 'CE',
                ))):
                candidates.append((r, c_idx, v_strip))
    # 키워드 매칭 실패 시: 최상단 짧은 후보 사용
    if candidates:
        candidates.sort(key=lambda x: (x[0], x[1]))
        return candidates[0][2]
    return None


# 기수 라벨: 제 N(당|전) 기 — 괄호·공백·키워드 다양성 포용
_PERIOD_LABEL_RE = re.compile(r'제\s*(\d+)\s*\(?\s*(당|전)\s*\)?\s*기')
# 날짜: YYYY[년/./-/ ]MM[월/./-/ ]DD
_DATE_RE = re.compile(
    r'(?<!\d)(20\d{2}|19\d{2})\s*[년.\-/]\s*(\d{1,2})\s*[월.\-/]\s*(\d{1,2})\s*일?'
)
# 통합: 라벨 + 날짜 한 셀 안 (40자 이내)
_PERIOD_FULL_RE = re.compile(
    r'제\s*(\d+)\s*\(?\s*(당|전)\s*\)?\s*기[^\n]{0,40}?'
    r'(20\d{2}|19\d{2})\s*[년.\-/]\s*(\d{1,2})\s*[월.\-/]\s*(\d{1,2})\s*일?'
)
# 키워드 + 날짜 (라벨 없이)
_PERIOD_KEYWORD_RE = re.compile(
    r'(당기|전기)\s*(?:말|초)?\s*[:：]?\s*'
    r'(20\d{2}|19\d{2})\s*[년.\-/]\s*(\d{1,2})\s*[월.\-/]\s*(\d{1,2})'
)


def _mk_period(fiscal_year, year, month, day):
    try:
        y, mo, d = int(year), int(month), int(day)
    except (TypeError, ValueError):
        return None
    if not (2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31):
        return None
    return {
        'fiscal_year': int(fiscal_year) if fiscal_year else None,
        'closing_year': y,
        'closing_date': f"{y:04d}-{mo:02d}-{d:02d}",
    }


def _extract_periods(ws):
    """첫 20행에서 period_current/period_prior 추출. 4단계 폴백."""
    period_current, period_prior = None, None
    labels = []   # (row, col, fy, kind)
    dates = []    # (row, col, year, month, day)

    # 1단계: 통합 정규식 (라벨 + 날짜 한 셀)
    for r in range(1, min(21, ws.max_row + 1)):
        for c_idx in range(1, min(16, ws.max_column + 1)):
            v = ws.cell(row=r, column=c_idx).value
            if not isinstance(v, str):
                continue
            for m in _PERIOD_FULL_RE.finditer(v):
                period_info = _mk_period(m.group(1), m.group(3), m.group(4), m.group(5))
                if not period_info:
                    continue
                if m.group(2) == '당' and not period_current:
                    period_current = period_info
                elif m.group(2) == '전' and not period_prior:
                    period_prior = period_info
            # 라벨·날짜 단독 수집 (폴백용)
            for lm in _PERIOD_LABEL_RE.finditer(v):
                labels.append((r, c_idx, lm.group(1), lm.group(2)))
            for dm in _DATE_RE.finditer(v):
                dates.append((r, c_idx, int(dm.group(1)), int(dm.group(2)), int(dm.group(3))))
            # 키워드 + 날짜 (당기/전기 ~ YYYY-MM-DD)
            for km in _PERIOD_KEYWORD_RE.finditer(v):
                period_info = _mk_period(None, km.group(2), km.group(3), km.group(4))
                if not period_info:
                    continue
                if km.group(1) == '당기' and not period_current:
                    period_current = period_info
                elif km.group(1) == '전기' and not period_prior:
                    period_prior = period_info

    # 2단계: 라벨·날짜 인접 매칭 (다른 셀 경계)
    if labels and dates and (not period_current or not period_prior):
        for lr, lc, fy, kind in labels:
            best, best_dist = None, 10**9
            for dr, dc, y, m, d in dates:
                dist = abs(dr - lr) * 100 + abs(dc - lc)
                if dist < best_dist:
                    best_dist, best = dist, (y, m, d)
            if best:
                period_info = _mk_period(fy, *best)
                if not period_info:
                    continue
                if kind == '당' and not period_current:
                    period_current = period_info
                elif kind == '전' and not period_prior:
                    period_prior = period_info

    # 3단계: 라벨도 없이 날짜만 있을 때 — 가장 최근 날짜를 당기로 추정
    if not period_current and dates:
        seen = set()
        unique = []
        for dr, dc, y, m, d in sorted(dates, key=lambda x: (-x[2], -x[3], -x[4])):
            key = (y, m, d)
            if key in seen:
                continue
            seen.add(key)
            unique.append((y, m, d))
        if unique:
            period_current = _mk_period(None, *unique[0])
            if len(unique) >= 2 and not period_prior:
                period_prior = _mk_period(None, *unique[1])

    return period_current, period_prior


def _extract_unit(ws):
    """단위(원/천원/백만원) 추출."""
    for r in range(1, min(16, ws.max_row + 1)):
        for c_idx in range(1, min(13, ws.max_column + 1)):
            v = ws.cell(row=r, column=c_idx).value
            if isinstance(v, str) and '단위' in v:
                if '백만원' in v:
                    return '백만원'
                if '천원' in v:
                    return '천원'
                if '원' in v:
                    return '원'
    return None


def dump_sheet_text(wb, sheet_name, max_rows=None, value_truncate=80):
    """시트의 모든 셀 값을 사람이 읽기 좋게 출력 (LLM이 시트 구조를 분석할 때 사용).
    
    Returns:
        문자열 - 행마다 'R{row}: A1=val | B1=val | ...' 형태
    """
    ws = wb[sheet_name]
    if max_rows is None:
        max_rows = ws.max_row
    lines = [f"## 시트: {sheet_name} ({ws.max_row}행 x {ws.max_column}열)"]
    for r in range(1, min(ws.max_row + 1, max_rows + 1)):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            col_letter = ws.cell(row=r, column=c).column_letter
            v_str = str(v)
            if len(v_str) > value_truncate:
                v_str = v_str[:value_truncate] + '...'
            cells.append(f"{col_letter}{r}={v_str}")
        if cells:
            lines.append(f"  {' | '.join(cells)}")
    return '\n'.join(lines)


# ─────────────────────── 검증 수식/서식 추가 ───────────────────────

def add_check(ws, log, name, cell_addr, formula, period='-', section='Section1'):
    """
    검증 수식 1건을 셀에 작성하고, 노란색 서식을 적용하고, 로그에 기록.
    
    Args:
        ws: openpyxl worksheet
        log: list (검증 로그 누적용)
        name: 검증 항목명 (예: 'BS 자산총계 당기')
        cell_addr: 수식이 들어갈 셀 (예: 'G68')
        formula: 검증 수식 (예: '=D68-(D11+D32)' - 정상이면 0)
        period: '당기' / '전기' / '-'
        section: 'Section1' (풋팅), 'Section1_Cross' (시트간 교차) 등
    
    검증 수식은 항상 "표시값 - 계산값" 형태로 작성한다 (정상이면 0, 오류면 차이값).
    """
    cell = ws[cell_addr]
    cell.value = formula
    cell.fill = YELLOW
    cell.font = NORMAL
    cell.border = BORDER
    cell.number_format = NUM_FMT
    cell.alignment = RIGHT
    log.append({
        'section': section,
        'sheet': ws.title,
        'name': name,
        'cell': cell_addr,
        'formula': formula,
        'period': period,
    })


def add_header(ws, cell_addr, text):
    """검증 컬럼 헤더 셀에 서식 적용 (예: G9 = '검증(당기)')"""
    c = ws[cell_addr]
    c.value = text
    c.font = BOLD
    c.fill = COL_HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER


def mark_narrative_error(ws, cell_addr, issue, fix, error_type='narrative'):
    """서술 오류 셀에 분홍색 + 셀메모 표시.
    
    Args:
        ws: worksheet
        cell_addr: 'B75' 등
        issue: 문제 설명 (메모에 표시)
        fix: 정정안
        error_type: 분류 태그 (요약시트에서 사용)
    """
    cell = ws[cell_addr]
    cell.fill = LIGHT_RED
    cell.font = ERR_FONT
    cell.comment = Comment(
        f"[{error_type}] {issue} | 정정안: {fix}", "Auditor"
    )


def disable_check_with_note(ws, cell_addr, note_text):
    """검증 불가 셀에 회색 안내 텍스트 표시 (텍스트형 셀 등)"""
    cell = ws[cell_addr]
    cell.value = note_text
    cell.fill = PatternFill('solid', start_color='D9D9D9', end_color='D9D9D9')
    cell.font = Font(name='맑은 고딕', size=9, italic=True, color='606060')
    cell.alignment = CENTER


# ─────────────────────── 재계산 + 오류 처리 ───────────────────────

def _find_soffice():
    """LibreOffice soffice 실행파일 경로 탐지. 못 찾으면 None.

    PATH 우선, 실패 시 OS별 표준 설치 경로 순회.
    """
    for cand in SOFFICE_CANDIDATES:
        # 절대경로면 존재 여부만, 아니면 PATH에서 검색
        if os.path.isabs(cand):
            if os.path.isfile(cand):
                return cand
        else:
            found = shutil.which(cand)
            if found:
                return found
    return None


def _libreoffice_recalc(path, timeout=60):
    """LibreOffice를 headless로 실행하여 수식을 재계산하고 같은 파일에 덮어쓴다.

    Args:
        path: 입력 xlsx 경로 (이 파일이 재계산 결과로 덮어써짐)
        timeout: soffice 실행 타임아웃(초)

    Raises:
        FileNotFoundError: soffice를 찾지 못했을 때
        RuntimeError: soffice가 비정상 종료하거나 출력 파일을 못 만들 때
        subprocess.TimeoutExpired: 타임아웃
    """
    soffice = _find_soffice()
    if soffice is None:
        raise FileNotFoundError(
            'LibreOffice(soffice) not found. Install LibreOffice and ensure '
            '"soffice" is in PATH, or install to a standard location. '
            f'Searched: {SOFFICE_CANDIDATES}'
        )

    src = Path(path).resolve()
    if not src.is_file():
        raise FileNotFoundError(f'Input file not found: {src}')

    # 동시 실행 충돌 방지를 위해 격리된 user profile + 출력 디렉토리 사용
    with tempfile.TemporaryDirectory(prefix='footing_recalc_') as tmp:
        tmp_path = Path(tmp)
        profile_dir = tmp_path / 'profile'
        out_dir = tmp_path / 'out'
        out_dir.mkdir()

        # convert-to xlsx 는 열고 재계산 후 저장하므로 수식 캐시값이 갱신됨
        proc = subprocess.run(
            [
                soffice,
                f'-env:UserInstallation=file:///{profile_dir.as_posix()}',
                '--headless',
                '--calc',
                '--convert-to', 'xlsx',
                '--outdir', str(out_dir),
                str(src),
            ],
            capture_output=True,
            timeout=timeout,
        )
        if proc.returncode != 0:
            stderr = proc.stderr.decode('utf-8', errors='replace')[-500:]
            raise RuntimeError(f'soffice failed (rc={proc.returncode}): {stderr}')

        out_file = out_dir / src.name
        if not out_file.is_file():
            # 다른 이름으로 저장된 경우 첫 xlsx를 찾는다
            xlsxs = list(out_dir.glob('*.xlsx'))
            if not xlsxs:
                raise RuntimeError(f'soffice produced no xlsx in {out_dir}')
            out_file = xlsxs[0]

        # 원본 위치로 덮어쓰기
        shutil.copyfile(str(out_file), str(src))


def _scan_errors(path):
    """재계산된 워크북을 열어 셀 단위로 오류값을 수집.

    Returns:
        {
          'status': 'success' | 'errors_found',
          'total_formulas': int,   # 캐시값이 있는 셀 수 (근사치)
          'total_errors':  int,
          'error_summary': {
              '#VALUE!': {'count': N, 'locations': ['Sheet!A1', ...]},
              ...
          }
        }
    """
    wb = load_workbook(path, data_only=True)
    error_summary = {}
    total_formulas = 0
    total_errors = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                # data_only=True 이므로 수식 셀은 캐시값으로 옴.
                # 숫자/문자/오류 모두 잡힘 — 오류만 추림.
                if isinstance(v, str) and v in ERROR_VALUES:
                    total_errors += 1
                    bucket = error_summary.setdefault(
                        v, {'count': 0, 'locations': []}
                    )
                    bucket['count'] += 1
                    if len(bucket['locations']) < 50:  # 너무 많아지면 잘라냄
                        bucket['locations'].append(f'{ws.title}!{cell.coordinate}')
                if isinstance(v, (int, float, str)):
                    total_formulas += 1
    wb.close()

    return {
        'status': 'success' if total_errors == 0 else 'errors_found',
        'total_formulas': total_formulas,
        'total_errors': total_errors,
        'error_summary': error_summary,
    }


def recalc_and_verify(path, max_attempts=3, timeout=60):
    """LibreOffice로 수식 재계산 후 오류값을 스캔. #VALUE! 자동 정리 후 재시도.

    Args:
        path: 워크북 경로 (재계산 결과로 덮어써짐)
        max_attempts: 자동 정리 후 최대 재시도 횟수
        timeout: 매 재계산 호출의 soffice 타임아웃(초)

    Returns:
        {'status': 'success'/'errors_found'/'recalc_failed',
         'total_errors': N, 'error_summary': {...}, 'total_formulas': N}
    """
    last_result = None
    for attempt in range(max_attempts):
        try:
            _libreoffice_recalc(path, timeout=timeout)
        except FileNotFoundError as exc:
            return {'status': 'recalc_failed', 'reason': str(exc)}
        except subprocess.TimeoutExpired:
            return {'status': 'recalc_failed', 'reason': f'timeout after {timeout}s'}
        except RuntimeError as exc:
            return {'status': 'recalc_failed', 'reason': str(exc)}

        data = _scan_errors(path)
        last_result = data
        if data['status'] == 'success':
            return data

        # 마지막 시도면 더 고치지 않고 반환
        if attempt >= max_attempts - 1:
            break

        fixed = _fix_value_errors(path, data['error_summary'])
        if fixed == 0:
            break  # 더 고칠 수 있는 게 없으면 중단

    return last_result


def _fix_value_errors(path, error_summary):
    """수식 텍스트로 표시된 셀(요약시트의 D열 등)이 #VALUE!를 내는 경우, 
    선두의 = 를 제거하여 일반 텍스트로 만든다.
    Returns: 고친 셀 수
    """
    wb = load_workbook(path)
    fixed_count = 0
    for err_type, info in error_summary.items():
        if err_type != '#VALUE!':
            continue
        for loc in info.get('locations', []):
            if '!' not in loc:
                continue
            sheet, cell = loc.split('!', 1)
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            try:
                v = ws[cell].value
            except (ValueError, KeyError):
                continue
            if isinstance(v, str) and v.startswith('=') and not v.upper().startswith('=HYPERLINK'):
                ws[cell] = v[1:]
                fixed_count += 1
    if fixed_count:
        wb.save(path)
    return fixed_count


# ─────────────────────── 검증결과요약 시트 생성 ───────────────────────

def build_summary_sheet(wb_path, log, narrative_findings, info,
                        section_titles=None, additional_summary_rows=None):
    """검증결과요약 시트를 워크북 맨 앞에 생성.
    
    Args:
        wb_path: 워크북 경로 (str/Path)
        log: list of check dicts (add_check 호출 결과 누적)
              각 dict: {'section', 'sheet', 'name', 'cell', 'formula', 'period'}
        narrative_findings: list of narrative finding dicts
              각 dict: {'type', 'sheet', 'cell', 'current', 'issue', 'fix'}
        info: analyze_workbook_basic() 결과 (회사명/기수/단위 등)
        section_titles: dict (선택) - {'inappropriate_reference': 'Section 2. 주석 참조 적절성', ...}
        additional_summary_rows: list (선택) - Section 6 종합 요약에 추가할 행
    """
    wb = load_workbook(str(wb_path))
    wb_d = load_workbook(str(wb_path), data_only=True)
    
    # 기존 요약 시트 제거
    for sht in ['검증결과요약', '교차검증_상세']:
        if sht in wb.sheetnames:
            del wb[sht]
    
    summary = wb.create_sheet('검증결과요약', 0)
    
    # 컬럼 너비
    widths = {'A': 6, 'B': 14, 'C': 30, 'D': 42, 'E': 16, 'F': 14, 'G': 22}
    for col, w in widths.items():
        summary.column_dimensions[col].width = w
    
    # 타이틀
    company = info.get('company', '회사명 미상')
    period = info.get('period_current') or {}
    fy = period.get('fiscal_year', '?')
    closing = period.get('closing_date', '?')
    unit = info.get('unit', '?')
    
    summary['A1'] = f'{company} - 제{fy}기 재무제표 풋팅 검증 결과'
    summary['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
    summary['A1'].fill = HEADER_FILL
    summary['A1'].alignment = CENTER
    summary.merge_cells('A1:G1')
    summary.row_dimensions[1].height = 35
    
    summary['A2'] = f'결산일: {closing} / 단위: {unit} / 검증일자: 자동생성'
    summary['A2'].font = Font(name='맑은 고딕', size=10, italic=True, color='606060')
    summary['A2'].alignment = CENTER
    summary.merge_cells('A2:G2')
    
    row = 4
    
    # ─── Section 1: 합계/소계 검증 (풋팅 + 교차검증) ───
    section1_items = [e for e in log if e['section'] == 'Section1']
    cross_items = [e for e in log if e['section'] == 'Section1_Cross']
    
    row = _write_section_header(summary, row, 'Section 1. 합계/소계 검증 (Footing)')
    section1_errors = 0
    if section1_items:
        row, errs = _write_check_table(summary, row, section1_items, wb_d)
        section1_errors = errs
    row = _write_section_summary(summary, row,
        f'Section 1 소결: 풋팅 {len(section1_items)}건 검증 / 오류 {section1_errors}건',
        section1_errors)
    row += 1
    
    if cross_items:
        row = _write_section_header(summary, row, 'Section 1-2. 시트간 교차검증 (Cross-footing)')
        row, cross_errors = _write_check_table(summary, row, cross_items, wb_d)
        row = _write_section_summary(summary, row,
            f'교차검증 소결: {len(cross_items)}건 / 오류 {cross_errors}건',
            cross_errors)
        row += 1
    else:
        cross_errors = 0
    
    # ─── Section 2~5: 서술 검증 ───
    default_titles = {
        'inappropriate_reference': 'Section 2. 주석 참조 적절성',
        'item_number': 'Section 3. 항목번호 순서',
        'period_inconsistency': 'Section 4. 당기/전기 연도 기재',
        'self_reference': 'Section 4-2. 자기참조',
        'narrative_value_mismatch': 'Section 5. 서술 내 수치 일치',
    }
    if section_titles:
        default_titles.update(section_titles)
    
    section_counts = {}
    for ftype, title in default_titles.items():
        items = [f for f in narrative_findings if f.get('type') == ftype]
        if not items:
            continue
        row = _write_section_header(summary, row, title)
        row = _write_findings_table(summary, row, items)
        section_counts[ftype] = len(items)
        row = _write_section_summary(summary, row,
            f'{title} 소결: {len(items)}건', len(items))
        row += 1
    
    # 분류되지 않은 finding은 별도로
    other_findings = [f for f in narrative_findings if f.get('type') not in default_titles]
    if other_findings:
        row = _write_section_header(summary, row, 'Section 5-기타. 기타 서술 발견사항')
        row = _write_findings_table(summary, row, other_findings)
        row = _write_section_summary(summary, row,
            f'기타 발견사항: {len(other_findings)}건', len(other_findings))
        row += 1
    
    # ─── Section 6: 종합 요약 ───
    row = _write_section_header(summary, row, 'Section 6. 종합 요약 (Summary of Findings)')
    headers6 = ['구분', '검증항목수', '오류건수', '주요내용']
    for i, h in enumerate(headers6):
        c = summary.cell(row=row, column=i+1, value=h)
        c.font = WHITE_BOLD_S
        c.fill = SUB_HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    summary.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
    row += 1
    
    summary_rows = [
        ('Section 1. 풋팅(합계/소계)',
         len(section1_items), section1_errors,
         '풋팅 모두 일치' if section1_errors == 0 else f'{section1_errors}건 차이 발견'),
    ]
    if cross_items:
        summary_rows.append((
            'Section 1-2. 교차검증',
            len(cross_items), cross_errors,
            '시트간 잔액 일치' if cross_errors == 0 else f'{cross_errors}건 불일치'
        ))
    for ftype, title in default_titles.items():
        cnt = section_counts.get(ftype, 0)
        if cnt > 0:
            summary_rows.append((title, cnt, cnt, ''))
    if additional_summary_rows:
        summary_rows.extend(additional_summary_rows)
    
    total_items = 0
    total_errors = 0
    for sect_name, total, errs, desc in summary_rows:
        summary.cell(row=row, column=1, value=sect_name).font = BOLD
        summary.cell(row=row, column=2, value=total).alignment = CENTER
        c_err = summary.cell(row=row, column=3, value=errs)
        c_err.alignment = CENTER
        if errs > 0:
            c_err.fill = RED
            c_err.font = WHITE_BOLD_S
        else:
            c_err.fill = LIGHT_GREEN
            c_err.font = BOLD
        summary.cell(row=row, column=4, value=desc).alignment = LEFT
        summary.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        for col in range(1, 8):
            summary.cell(row=row, column=col).border = BORDER
        total_items += total
        total_errors += errs
        row += 1
    
    # 합계 행
    summary.cell(row=row, column=1, value='합  계').font = WHITE_BOLD_S
    summary.cell(row=row, column=1).fill = HEADER_FILL
    summary.cell(row=row, column=1).alignment = CENTER
    summary.cell(row=row, column=2, value=total_items).font = WHITE_BOLD_S
    summary.cell(row=row, column=2).fill = HEADER_FILL
    summary.cell(row=row, column=2).alignment = CENTER
    c_t = summary.cell(row=row, column=3, value=total_errors)
    c_t.font = WHITE_BOLD_S
    c_t.fill = RED if total_errors > 0 else LIGHT_GREEN
    c_t.alignment = CENTER
    msg = '모든 검증 통과' if total_errors == 0 else f'총 {total_errors}건 정정 필요'
    summary.cell(row=row, column=4, value=msg).font = WHITE_BOLD_S
    summary.cell(row=row, column=4).fill = HEADER_FILL
    summary.cell(row=row, column=4).alignment = LEFT
    summary.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
    for col in range(1, 8):
        summary.cell(row=row, column=col).border = BORDER
    
    summary.freeze_panes = 'A4'
    wb.save(str(wb_path))
    return {'total_items': total_items, 'total_errors': total_errors}


def _write_section_header(ws, row, title):
    c = ws.cell(row=row, column=1, value=title)
    c.font = WHITE_BOLD
    c.fill = HEADER_FILL
    c.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.row_dimensions[row].height = 25
    return row + 1


def _write_check_table(ws, row, items, wb_d):
    """검증 항목 테이블. Returns (next_row, error_count)"""
    headers = ['No', '시트', '검증항목', '수식', '결과', '판정', '바로가기']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = WHITE_BOLD_S
        c.fill = SUB_HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    row += 1
    
    error_count = 0
    for i, e in enumerate(items, 1):
        try:
            val = wb_d[e['sheet']][e['cell']].value
        except Exception:
            val = None
        diff = val if isinstance(val, (int, float)) else 0
        is_err = (diff != 0) and val is not None
        if is_err:
            error_count += 1
        
        ws.cell(row=row, column=1, value=i).alignment = CENTER
        ws.cell(row=row, column=2, value=e['sheet']).alignment = CENTER
        ws.cell(row=row, column=3, value=e['name']).alignment = LEFT
        ftext = e['formula'][1:] if e['formula'].startswith('=') else e['formula']
        c_f = ws.cell(row=row, column=4, value=ftext)
        c_f.font = MONO
        c_f.alignment = LEFT
        c_r = ws.cell(row=row, column=5, value=diff)
        c_r.number_format = NUM_FMT
        c_r.alignment = RIGHT
        c_j = ws.cell(row=row, column=6, value='OK' if not is_err else '오류')
        c_j.alignment = CENTER
        if is_err:
            c_j.fill = RED
            c_j.font = WHITE_BOLD_S
            c_r.fill = RED
            c_r.font = WHITE_BOLD_S
        else:
            c_j.fill = LIGHT_GREEN
            c_j.font = BOLD
        c_l = ws.cell(row=row, column=7,
                      value=hyperlink(e['sheet'], e['cell'], f"{e['sheet']}!{e['cell']}"))
        c_l.font = LINK
        c_l.alignment = CENTER
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
            cur = ws.cell(row=row, column=col).font
            if col not in [4, 5, 6, 7] and not cur.bold:
                ws.cell(row=row, column=col).font = NORMAL
        row += 1
    return row, error_count


def _write_findings_table(ws, row, findings):
    """서술 발견사항 테이블"""
    headers = ['No', '시트', '셀', '현재 표기', '문제', '정정안', '바로가기']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = WHITE_BOLD_S
        c.fill = SUB_HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    row += 1
    
    for i, f in enumerate(findings, 1):
        ws.cell(row=row, column=1, value=i).alignment = CENTER
        ws.cell(row=row, column=2, value=f['sheet']).alignment = CENTER
        ws.cell(row=row, column=3, value=f['cell']).alignment = CENTER
        ws.cell(row=row, column=4, value=str(f.get('current', ''))[:80]).alignment = LEFT
        c_i = ws.cell(row=row, column=5, value=f.get('issue', ''))
        c_i.alignment = LEFT
        c_i.fill = LIGHT_RED
        c_i.font = ERR_FONT
        c_x = ws.cell(row=row, column=6, value=str(f.get('fix', ''))[:80])
        c_x.alignment = LEFT
        c_x.fill = LIGHT_GREEN
        c_x.font = BOLD
        c_l = ws.cell(row=row, column=7,
                      value=hyperlink(f['sheet'], f['cell'], f"{f['sheet']}!{f['cell']}"))
        c_l.font = LINK
        c_l.alignment = CENTER
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
    return row


def _write_section_summary(ws, row, text, error_count):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD
    c.fill = LIGHT_RED if error_count > 0 else LIGHT_GREEN
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    return row + 1
