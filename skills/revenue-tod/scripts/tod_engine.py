"""
매출 TOD 엔진 - 범용 템플릿
============================
이 스크립트는 매출 Test of Details의 핵심 로직을 담고 있다.
클라이언트별로 수정이 필요한 부분:
  1. SAMPLE_EXTRACTION 섹션: 엑셀 구조(시트명, 행/열 범위)에 맞게 수정
  2. get_folder_path(): 폴더 구조에 맞게 수정
  3. BASE 경로: 실제 증빙 폴더 경로로 변경

사용법:
  1. 아래 "=== 클라이언트별 수정 영역 ===" 섹션을 수정
  2. python3 tod_engine.py 실행
  3. tod_detailed_results.json 생성됨 → tod_workpaper.py로 엑셀 변환
"""

import pdfplumber
import openpyxl
import os
import re
import json
from pathlib import Path


# ============================================================
# === 클라이언트별 수정 영역 ===
# ============================================================

# 증빙 폴더 최상위 경로
BASE = "/path/to/evidence/folder"

# 결과 저장 경로
OUTPUT_JSON = "/path/to/tod_detailed_results.json"

# ============================================================


# ========================
# 공통 유틸리티 함수
# ========================

def extract_pdf_text(filepath):
    """PDF에서 텍스트 추출. 실패 시 에러 메시지 반환."""
    try:
        with pdfplumber.open(filepath) as pdf:
            text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
            return text.strip()
    except Exception as e:
        return f"[PDF READ ERROR: {e}]"


def extract_numbers(text):
    """텍스트에서 숫자(금액) 추출"""
    numbers = re.findall(r'[\d,]+\.?\d*', text)
    result = []
    for n in numbers:
        try:
            val = float(n.replace(',', ''))
            if val > 0:
                result.append(val)
        except:
            pass
    return result


def find_amount_in_text(text, target_amount, tolerance=1.0):
    """텍스트 내 목표 금액 존재 여부 확인. (found, matched_value) 반환."""
    numbers = extract_numbers(text)
    for n in numbers:
        if abs(n - target_amount) < tolerance:
            return True, n
    return False, None


def find_string_in_text(text, target, partial=True):
    """텍스트 내 목표 문자열 존재 여부 (부분매칭 포함)"""
    if not target or not text:
        return False
    if target.lower() in text.lower():
        return True
    if partial:
        parts = target.strip().split()
        for p in parts:
            if len(p) > 2 and p.lower() in text.lower():
                return True
    return False


def extract_incoterms(text):
    """텍스트에서 Incoterms 추출"""
    terms = ['DAP', 'DDP', 'DAT', 'DPU', 'EXW', 'FOB', 'FCA', 'CFR', 'CIF', 'CIP', 'CPT', 'FAS']
    for t in terms:
        if re.search(rf'\b{t}\b', text, re.IGNORECASE):
            return t
    return None


def extract_dates(text):
    """텍스트에서 날짜 패턴 추출"""
    patterns = [
        r'\d{4}[-/.]\d{1,2}[-/.]\d{1,2}',
        r'\d{1,2}[-/.]\d{1,2}[-/.]\d{4}',
    ]
    dates = []
    for p in patterns:
        dates.extend(re.findall(p, text))
    return dates


# ========================
# 파일 분류 (우선순위 중요!)
# ========================

def classify_file(filename):
    """
    파일명 기반 증빙 분류.
    순서: 세금계산서 → 거래명세서 → POD → BL → CI → PO → INV/PL → OTHER
    """
    fn = filename.lower()

    # 1. 세금계산서 (최우선)
    if '세금계산서' in fn:
        return '세금계산서'

    # 2. 거래명세서/인수증
    if any(x in fn for x in ['거래명세', '인수증']):
        return '거래명세서'

    # 3. POD
    if any(x in fn for x in ['pod', 'proof of delivery', 'delivery receipt',
                               '배송완료', '배송 서명', 'delivery order', 'pickup proof']):
        return 'POD'

    # 4. BL/AWB/출고증
    if any(x in fn for x in ['bl ', 'bl-', 'bl_', 'bl#', 'b/l', 'hawb', 'hbl', 'awb',
                               'air waybill', 'airwaybill', 'waybill', 'transportlabel',
                               'transport label', 'fedex', 'dhl', 'ups -', 'ups_',
                               '선적', 'skor', 'sur bl', 'sur_bl', 'sur ',
                               '출고증', '출고관련서류', '출고 증', 'kbe2', 'kbl0']):
        if 'pod' not in fn:
            return 'BL'

    # 5. CI (Commercial Invoice) - 단독 CI만
    if any(x in fn for x in ['ci-', 'ci_', 'commercial invoice']):
        if 'cipl' not in fn and 'ci&pl' not in fn:
            return 'CI'

    # 6. PO (Purchase Order)
    if any(x in fn for x in ['po ', 'po-', 'po_', 'po.', 'purchase order',
                               'p_us_', 'p_my_', 'p_uk_', 'p_de_', 'p_pt_', 'p_kor_',
                               'poimp', 'poinmt', 'pi -', 'pi_']):
        if 'pod' not in fn and 'invoice' not in fn and '출고' not in fn:
            return 'PO'

    # 7. INV/PL (Invoice/Packing List, 복합서류)
    if any(x in fn for x in ['inv_', 'inv-', 'inv ', 'inv0', 'invoice',
                               'cipl', 'ci&pl', 'pl&inv', '출고완료', '출고 완료']):
        return 'INV/PL'

    # 8. Order류
    if 'order' in fn and 'delivery' not in fn:
        return 'PO'

    return 'OTHER'


# ========================
# 폴더 경로 결정
# ========================

def get_folder_path(sample):
    """
    샘플의 매출유형과 폴더명에 따라 실제 폴더 경로를 반환.
    ** 클라이언트별 폴더 구조에 맞게 수정 필요 **
    """
    folder = sample['folder']
    stype = sample['type']

    if stype == '국내매출':
        return os.path.join(BASE, '국내매출', folder)
    elif stype == '중계무역':
        sub = '26년분' if '26년' in folder else '25년분'
        return os.path.join(BASE, '중계무역', sub, folder)
    elif stype == '직수출':
        sub = '26년분' if '26년' in folder else '25년분'
        return os.path.join(BASE, '직수출', sub, folder)
    return None


# ========================
# 해외매출 분석 (6개 테스트)
# ========================

def analyze_overseas_sample(sample):
    """해외매출 샘플 상세 분석 - 모든 증빙 서류를 읽고 교차 대사"""
    result = {
        'folder': sample['folder'],
        'type': sample['type'],
        'ci_no_excel': sample.get('ci_no', ''),
        'customer_excel': sample.get('customer', ''),
        'incoterms_excel': sample.get('incoterms', ''),
        'currency_excel': sample.get('currency', ''),
        'fc_amount_excel': sample.get('fc_amount'),
        'krw_amount_excel': sample.get('krw_amount'),
        'ci_date_excel': sample.get('ci_date', ''),
        'period': sample.get('period', ''),
        'docs': {},
        'test_1_po_order': {'result': '', 'detail': ''},
        'test_2_invoice': {'result': '', 'detail': ''},
        'test_3_shipping': {'result': '', 'detail': ''},
        'test_4_incoterms_revenue': {'result': '', 'detail': ''},
        'test_5_pod_arrival': {'result': '', 'detail': ''},
        'test_6_amount_match': {'result': '', 'detail': ''},
        'overall': 'Pass',
        'exceptions': [],
    }

    folder_path = get_folder_path(sample)
    if not folder_path or not os.path.exists(folder_path):
        result['overall'] = 'Fail'
        result['exceptions'].append('폴더 미존재')
        for t in ['test_1_po_order', 'test_2_invoice', 'test_3_shipping',
                   'test_4_incoterms_revenue', 'test_5_pod_arrival', 'test_6_amount_match']:
            result[t] = {'result': 'Fail', 'detail': '증빙 폴더 미존재'}
        return result

    # --- 파일 읽기 및 분류 ---
    files = sorted([f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))])
    ci_docs, po_docs, bl_docs, pod_docs, inv_pl_docs, other_docs = [], [], [], [], [], []

    for f in files:
        fpath = os.path.join(folder_path, f)
        ftype = classify_file(f)
        ext = f.lower().split('.')[-1]
        doc_info = {'filename': f, 'type': ftype, 'text': '', 'readable': False}

        if ext in ['pdf']:
            text = extract_pdf_text(fpath)
            if '[PDF READ ERROR' not in text and text.strip():
                doc_info['text'] = text
                doc_info['readable'] = True
            else:
                doc_info['text'] = '[텍스트 추출 불가]'
        elif ext in ['jpg', 'jpeg', 'png']:
            doc_info['text'] = '[이미지 파일 - 텍스트 추출 불가]'
        else:
            doc_info['text'] = '[지원하지 않는 형식]'

        if ftype == 'CI': ci_docs.append(doc_info)
        elif ftype == 'PO': po_docs.append(doc_info)
        elif ftype == 'BL': bl_docs.append(doc_info)
        elif ftype == 'POD': pod_docs.append(doc_info)
        elif ftype == 'INV/PL': inv_pl_docs.append(doc_info)
        else: other_docs.append(doc_info)

    result['docs'] = {
        'CI': ci_docs, 'PO': po_docs, 'BL': bl_docs,
        'POD': pod_docs, 'INV/PL': inv_pl_docs, 'OTHER': other_docs
    }

    # ===== T1: PO - 고객 주문 확인 =====
    if po_docs:
        po_names = [d['filename'] for d in po_docs]
        detail_parts = [f"PO 파일 확인: {', '.join(po_names)}"]
        cust_found = False
        for d in po_docs:
            if d['readable'] and find_string_in_text(d['text'], sample['customer']):
                cust_found = True
                detail_parts.append(f"PO({d['filename']})에서 거래처명 '{sample['customer']}' 확인")
                break
        if not cust_found:
            if po_docs[0]['readable']:
                detail_parts.append("PO 텍스트에서 거래처명 자동매칭 불가 - 수동확인 필요")
            else:
                detail_parts.append("PO 텍스트 추출 불가 - 수동확인 필요")
        result['test_1_po_order'] = {'result': 'Pass', 'detail': ' / '.join(detail_parts)}
    else:
        result['test_1_po_order'] = {'result': 'Fail', 'detail': 'PO 서류 미확인 - 고객 주문 증빙 없음'}
        result['exceptions'].append('PO 미확인')

    # ===== T2: Invoice(CI) - 빌링 확인 =====
    invoice_doc = ci_docs[0] if ci_docs else None
    if not invoice_doc and inv_pl_docs:
        for d in inv_pl_docs:
            if d['readable']:
                invoice_doc = d
                break

    if invoice_doc and invoice_doc['readable']:
        inv_text = invoice_doc['text']
        detail_parts = [f"Invoice 파일: {invoice_doc['filename']}"]

        ci_no = sample.get('ci_no', '')
        ci_base = ci_no.split('(')[0].strip() if ci_no else ''
        if ci_base and ci_base.lower() in inv_text.lower():
            detail_parts.append(f"CI번호 '{ci_base}' → Invoice 상 확인됨 [일치]")
        elif ci_base:
            ci_short = re.sub(r'\(.*?\)', '', ci_base).strip()
            if ci_short.lower() in invoice_doc['filename'].lower():
                detail_parts.append(f"CI번호 '{ci_base}' → 파일명에서 확인됨 [일치]")
            else:
                detail_parts.append(f"CI번호 '{ci_base}' → 자동매칭 불가 [수동확인 필요]")

        if find_string_in_text(inv_text, sample['customer']):
            detail_parts.append(f"거래처 '{sample['customer']}' → Invoice 상 확인됨 [일치]")
        else:
            detail_parts.append(f"거래처 '{sample['customer']}' → Invoice 텍스트 자동매칭 불가 [수동확인 필요]")
            result['exceptions'].append('Invoice 거래처 자동대조 불가')

        result['test_2_invoice'] = {'result': 'Pass', 'detail': ' / '.join(detail_parts)}
    elif invoice_doc and not invoice_doc['readable']:
        result['test_2_invoice'] = {
            'result': 'Exception',
            'detail': f"Invoice 파일: {invoice_doc['filename']} - 텍스트 추출 불가, 수동확인 필요"
        }
        result['exceptions'].append('Invoice 텍스트 추출 불가')
    else:
        result['test_2_invoice'] = {'result': 'Fail', 'detail': 'Invoice(CI) 서류 미확인'}
        result['exceptions'].append('Invoice(CI) 미확인')

    # ===== T3: 선적서류 - 인도 확인 =====
    ship_docs = bl_docs + [d for d in inv_pl_docs if '출고' in d['filename']]
    if ship_docs:
        ship_names = [d['filename'] for d in ship_docs]
        detail_parts = [f"선적/출고서류 확인: {', '.join(ship_names)}"]
        for d in ship_docs:
            if d['readable']:
                dates = extract_dates(d['text'])
                if dates:
                    detail_parts.append(f"{d['filename']} 내 일자 확인: {dates[0]}")
                    break
        result['test_3_shipping'] = {'result': 'Pass', 'detail': ' / '.join(detail_parts)}
    else:
        result['test_3_shipping'] = {'result': 'Fail', 'detail': '선적서류(BL/AWB) 또는 출고증 미확인'}
        result['exceptions'].append('선적/출고서류 미확인')

    # ===== T4: Incoterms & 매출인식 =====
    incoterms_excel = sample.get('incoterms', '')
    incoterms_on_doc = None
    detail_parts = [f"엑셀상 운임조건: {incoterms_excel}"]

    if invoice_doc and invoice_doc['readable']:
        incoterms_on_doc = extract_incoterms(invoice_doc['text'])
        if incoterms_on_doc:
            if incoterms_on_doc.upper() == incoterms_excel.upper():
                detail_parts.append(f"Invoice상 운임조건: {incoterms_on_doc} [일치]")
            else:
                detail_parts.append(f"Invoice상 운임조건: {incoterms_on_doc} [불일치 - 엑셀: {incoterms_excel}]")
                result['exceptions'].append(f'운임조건 불일치 (엑셀: {incoterms_excel}, Invoice: {incoterms_on_doc})')
        else:
            detail_parts.append("Invoice에서 운임조건 자동추출 불가 - 수동확인 필요")

    effective_terms = incoterms_on_doc or incoterms_excel
    if effective_terms in ['EXW', 'FCA', 'FOB', 'FAS']:
        detail_parts.append(f"{effective_terms} 조건: 출고/선적 시점 매출인식 → 출고증/BL 확인으로 충족")
        if ship_docs or bl_docs:
            detail_parts.append("매출인식조건 충족 확인")
        else:
            detail_parts.append("출고/선적 증빙 미확인 - 매출인식조건 충족 여부 수동확인 필요")
    elif effective_terms in ['CFR', 'CIF', 'CIP', 'CPT']:
        detail_parts.append(f"{effective_terms} 조건: 선적 시점 매출인식 → BL/AWB 확인으로 충족")
        if bl_docs:
            detail_parts.append("매출인식조건 충족 확인")
        else:
            detail_parts.append("선적서류 미확인 - 매출인식조건 충족 여부 수동확인 필요")
    elif effective_terms in ['DAP', 'DDP', 'DAT', 'DPU']:
        detail_parts.append(f"{effective_terms} 조건: 도착지 인도 시점 매출인식 → POD 확인 필요")
        if pod_docs:
            detail_parts.append("POD 확인 → 매출인식조건 충족")
        else:
            detail_parts.append("POD 미확인 - 매출인식조건 충족 여부 수동확인 필요")

    result['test_4_incoterms_revenue'] = {'result': 'Pass', 'detail': ' / '.join(detail_parts)}

    # ===== T5: POD 도착 확인 (D조건) =====
    d_terms = ['DAP', 'DDP', 'DAT', 'DPU']
    if effective_terms in d_terms:
        if pod_docs:
            pod_names = [d['filename'] for d in pod_docs]
            pod_detail = f"POD 확인: {', '.join(pod_names)}"
            for d in pod_docs:
                if d['readable']:
                    dates = extract_dates(d['text'])
                    if dates:
                        pod_detail += f" / 배송완료일: {dates[0]}"
                        break
            result['test_5_pod_arrival'] = {'result': 'Pass', 'detail': pod_detail}
        else:
            result['test_5_pod_arrival'] = {
                'result': 'Exception',
                'detail': f'{effective_terms} 조건이나 POD(도착증빙) 미확인 - 매출인식시점 수동확인 필요'
            }
            result['exceptions'].append(f'POD 미확인 ({effective_terms} 조건)')
    else:
        result['test_5_pod_arrival'] = {
            'result': 'N/A',
            'detail': f'{effective_terms} 조건 - 도착증빙 불필요 (선적/출고 시점 매출인식)'
        }

    # ===== T6: 금액 대조 =====
    fc_amount = sample.get('fc_amount')
    currency = sample.get('currency', '')
    krw_amount = sample.get('krw_amount')
    amt_parts = []

    if fc_amount:
        amt_parts.append(f"엑셀상 금액: {currency} {fc_amount:,.2f} (KRW {krw_amount:,.0f})")
    else:
        amt_parts.append(f"엑셀상 금액: KRW {krw_amount:,.0f}")

    amount_matched = False
    if invoice_doc and invoice_doc['readable'] and fc_amount:
        found, matched_val = find_amount_in_text(invoice_doc['text'], fc_amount)
        if found:
            amt_parts.append(f"Invoice({invoice_doc['filename']})상 금액: {currency} {matched_val:,.2f} → 엑셀 {currency} {fc_amount:,.2f} [일치]")
            amount_matched = True
        else:
            found_krw, matched_krw = find_amount_in_text(invoice_doc['text'], krw_amount)
            if found_krw:
                amt_parts.append(f"Invoice상 KRW 금액: {matched_krw:,.0f} → 엑셀 KRW {krw_amount:,.0f} [일치]")
                amount_matched = True
            else:
                amt_parts.append(f"Invoice 텍스트에서 금액 자동매칭 불가 ({currency} {fc_amount:,.2f}) [수동확인 필요]")
                result['exceptions'].append('Invoice 금액 자동대조 불가')
    elif invoice_doc:
        amt_parts.append("Invoice 텍스트 추출 불가 - 금액 수동확인 필요")
        result['exceptions'].append('Invoice 금액 확인 불가 (텍스트 추출 실패)')
    else:
        amt_parts.append("Invoice 미확인으로 금액 대조 불가")

    result['test_6_amount_match'] = {
        'result': 'Pass' if amount_matched else 'Exception',
        'detail': ' / '.join(amt_parts)
    }

    # --- 종합 판정 ---
    if any(t['result'] == 'Fail' for t in [result['test_1_po_order'], result['test_2_invoice'], result['test_3_shipping']]):
        result['overall'] = 'Fail'
    elif result['exceptions']:
        result['overall'] = 'Exception'
    else:
        result['overall'] = 'Pass'

    return result


# ========================
# 국내매출 분석
# ========================

def analyze_domestic_sample(sample):
    """국내매출 샘플 상세 분석"""
    result = {
        'folder': sample['folder'],
        'type': sample['type'],
        'customer_excel': sample.get('customer', ''),
        'krw_amount_excel': sample.get('krw_amount'),
        'account_date_excel': sample.get('account_date', ''),
        'voucher_no': sample.get('voucher_no', ''),
        'docs': {},
        'test_tax_invoice': {'result': '', 'detail': ''},
        'test_transaction_stmt': {'result': '', 'detail': ''},
        'test_amount_match': {'result': '', 'detail': ''},
        'overall': 'Pass',
        'exceptions': [],
    }

    folder_path = get_folder_path(sample)
    if not folder_path or not os.path.exists(folder_path):
        result['overall'] = 'Fail'
        result['exceptions'].append('폴더 미존재')
        return result

    files = sorted([f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))])
    tax_inv_docs, trans_stmt_docs, other_docs = [], [], []

    for f in files:
        fpath = os.path.join(folder_path, f)
        ftype = classify_file(f)
        doc_info = {'filename': f, 'type': ftype, 'text': '', 'readable': False}
        ext = f.lower().split('.')[-1]

        if ext == 'pdf':
            text = extract_pdf_text(fpath)
            if '[PDF READ ERROR' not in text and text.strip():
                doc_info['text'] = text
                doc_info['readable'] = True
            else:
                doc_info['text'] = '[텍스트 추출 불가]'
        else:
            doc_info['text'] = '[지원하지 않는 형식]'

        if ftype == '세금계산서': tax_inv_docs.append(doc_info)
        elif ftype == '거래명세서': trans_stmt_docs.append(doc_info)
        else: other_docs.append(doc_info)

    result['docs'] = {'세금계산서': tax_inv_docs, '거래명세서': trans_stmt_docs, 'OTHER': other_docs}

    krw_amt = sample.get('krw_amount', 0) or 0

    # ① 세금계산서
    if tax_inv_docs:
        d = tax_inv_docs[0]
        detail_parts = [f"세금계산서 파일: {d['filename']}"]
        if d['readable']:
            found, matched = find_amount_in_text(d['text'], krw_amt)
            if found:
                detail_parts.append(f"공급가액: {matched:,.0f}원 → 엑셀 {krw_amt:,.0f}원 [일치]")
            else:
                detail_parts.append(f"공급가액 자동매칭 불가 (엑셀: {krw_amt:,.0f}원) [수동확인 필요]")
                result['exceptions'].append('세금계산서 금액 자동대조 불가')

            cust = sample.get('customer', '')
            cust_variants = [cust]
            if '(주)' in cust:
                cust_variants.append(cust.replace('(주)', '').strip())
                cust_variants.append(cust.replace('(주)', '주식회사').strip())
                cust_variants.append('주식회사 ' + cust.replace('(주)', '').strip())
            if '주식회사' in cust:
                cust_variants.append(cust.replace('주식회사 ', '').strip())

            cust_found = False
            for cv in cust_variants:
                if cv in d['text']:
                    detail_parts.append(f"거래처: '{cv}' → 세금계산서 상 확인됨 [일치]")
                    cust_found = True
                    break
            if not cust_found:
                detail_parts.append(f"거래처 '{cust}' 자동매칭 불가 [수동확인 필요]")
                result['exceptions'].append('세금계산서 거래처 자동대조 불가')

            dates = extract_dates(d['text'])
            if dates:
                detail_parts.append(f"작성일자: {dates[0]}")
        else:
            detail_parts.append("텍스트 추출 불가 - 수동확인 필요")
            result['exceptions'].append('세금계산서 텍스트 추출 불가')

        has_exc = any('세금계산서' in e for e in result['exceptions'])
        result['test_tax_invoice'] = {
            'result': 'Exception' if has_exc else 'Pass',
            'detail': ' / '.join(detail_parts)
        }
    else:
        result['test_tax_invoice'] = {'result': 'Fail', 'detail': '세금계산서 미확인'}
        result['exceptions'].append('세금계산서 미확인')

    # ② 거래명세서
    if trans_stmt_docs:
        d = trans_stmt_docs[0]
        detail_parts = [f"거래명세서 파일: {d['filename']}"]
        if d['readable']:
            found, matched = find_amount_in_text(d['text'], krw_amt)
            if found:
                detail_parts.append(f"합계금액: {matched:,.0f}원 → 엑셀 {krw_amt:,.0f}원 [일치]")
            else:
                detail_parts.append("금액 자동매칭 불가 [수동확인 필요]")
        else:
            detail_parts.append("텍스트 추출 불가 - 수동확인 필요")
        result['test_transaction_stmt'] = {'result': 'Pass', 'detail': ' / '.join(detail_parts)}
    else:
        result['test_transaction_stmt'] = {'result': 'Exception', 'detail': '거래명세서(인수증) 미확인 - 수동확인 필요'}
        result['exceptions'].append('거래명세서 미확인')

    # ③ 금액 종합
    amt_parts = [f"엑셀상 매출액(대변): {krw_amt:,.0f}원"]
    if tax_inv_docs and tax_inv_docs[0]['readable']:
        found, matched = find_amount_in_text(tax_inv_docs[0]['text'], krw_amt)
        if found:
            amt_parts.append(f"세금계산서 공급가액: {matched:,.0f}원 [일치]")
    if trans_stmt_docs and trans_stmt_docs[0]['readable']:
        found, matched = find_amount_in_text(trans_stmt_docs[0]['text'], krw_amt)
        if found:
            amt_parts.append(f"거래명세서 합계: {matched:,.0f}원 [일치]")

    result['test_amount_match'] = {
        'result': 'Pass' if '일치' in ' '.join(amt_parts) else 'Exception',
        'detail': ' / '.join(amt_parts)
    }

    if result['exceptions']:
        result['overall'] = 'Exception'

    return result


# ============================================================
# === 샘플 추출 (SAMPLE_EXTRACTION) ===
# === 클라이언트별로 이 부분을 수정한다 ===
# ============================================================

def extract_samples_from_excel(excel_path):
    """
    엑셀에서 샘플 데이터 추출.
    ** 이 함수는 클라이언트별 엑셀 구조에 맞게 수정해야 한다. **

    반환값: list of dict, 각 dict 필수키:
      - folder: 폴더명 (str)
      - type: 매출유형 ('국내매출', '중계무역', '직수출' 등)
      - customer: 거래처명 (str)
      - krw_amount: 원화금액 (float)
    해외매출 추가키:
      - ci_no, ci_date, incoterms, currency, fc_amount
    국내매출 추가키:
      - account_date, voucher_no
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active  # 또는 wb['시트명']
    samples = []

    # === 여기에 클라이언트별 행/열 범위에 맞는 추출 로직 작성 ===
    # 예시:
    # for r in range(start_row, end_row + 1):
    #     folder = ws.cell(r, folder_col).value
    #     if folder:
    #         samples.append({
    #             'folder': str(folder),
    #             'type': '직수출',
    #             'ci_no': str(ws.cell(r, ci_col).value or ''),
    #             'customer': str(ws.cell(r, cust_col).value or ''),
    #             'incoterms': str(ws.cell(r, inco_col).value or ''),
    #             'krw_amount': ws.cell(r, krw_col).value,
    #             'currency': str(ws.cell(r, curr_col).value or ''),
    #             'fc_amount': ws.cell(r, fc_col).value,
    #         })

    print(f"[WARNING] extract_samples_from_excel()이 수정되지 않았습니다.")
    print(f"클라이언트 엑셀 구조에 맞게 이 함수를 수정하세요.")

    return samples


# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    # 엑셀 경로 - 클라이언트별 수정
    excel_path = os.path.join(BASE, "샘플링리스트.xlsx")

    samples = extract_samples_from_excel(excel_path)
    print(f"샘플 {len(samples)}건 추출 완료")

    all_results = []
    for i, sample in enumerate(samples):
        print(f"[{i+1}/{len(samples)}] {sample['folder']}...", end=' ')
        if sample['type'] == '국내매출':
            r = analyze_domestic_sample(sample)
        else:
            r = analyze_overseas_sample(sample)
        all_results.append(r)
        print(f"{r['overall']} ({len(r['exceptions'])} exceptions)")

    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2, default=str)

    p = sum(1 for r in all_results if r['overall'] == 'Pass')
    e = sum(1 for r in all_results if r['overall'] == 'Exception')
    fail = sum(1 for r in all_results if r['overall'] == 'Fail')
    print(f"\n=== SUMMARY: Pass={p}, Exception={e}, Fail={fail} (Total={len(all_results)}) ===")
    print(f"결과 저장: {OUTPUT_JSON}")
