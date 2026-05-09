"""
매출 TOD 워크페이퍼 생성기 - 범용 템플릿
==========================================
tod_engine.py의 결과(JSON)를 읽어 3시트 엑셀 워크페이퍼를 생성한다.

사용법:
  python3 tod_workpaper.py <results.json> <output.xlsx> [회사명] [감사기간]

예시:
  python3 tod_workpaper.py tod_detailed_results.json 매출TOD_워크페이퍼.xlsx "(주)삼성전자" "2025년 1월~2026년 1월"
"""

import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ========================
# 스타일 정의
# ========================

header_font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
pass_fill = PatternFill('solid', fgColor='C6EFCE')
fail_fill = PatternFill('solid', fgColor='FFC7CE')
exception_fill = PatternFill('solid', fgColor='FFEB9C')
na_fill = PatternFill('solid', fgColor='D9E2F3')
title_font = Font(name='맑은 고딕', bold=True, size=14)
subtitle_font = Font(name='맑은 고딕', bold=True, size=11)
data_font = Font(name='맑은 고딕', size=8)
data_font_bold = Font(name='맑은 고딕', size=8, bold=True)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap = Alignment(vertical='top', wrap_text=True)
wrap_center = Alignment(horizontal='center', vertical='top', wrap_text=True)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border


def style_data(ws, row, max_col, font=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or data_font
        cell.border = border
        cell.alignment = wrap


def color_result(cell, value):
    cell.alignment = center
    if value == 'Pass':
        cell.fill = pass_fill
    elif value == 'Exception':
        cell.fill = exception_fill
    elif value == 'Fail':
        cell.fill = fail_fill
    elif value == 'N/A':
        cell.fill = na_fill


def generate_workpaper(results, output_path, company_name="", audit_period=""):
    wb = openpyxl.Workbook()

    # ============ Sheet 1: Summary ============
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    ws_sum['A1'] = f'매출 TOD 워크페이퍼 - {company_name}'
    ws_sum['A1'].font = title_font
    ws_sum['A2'] = f'감사대상기간: {audit_period}'
    ws_sum['A2'].font = subtitle_font

    ws_sum['A5'] = '1. 테스트 요약'
    ws_sum['A5'].font = subtitle_font

    headers = ['구분', '총 샘플수', 'Pass', 'Exception', 'Fail', 'Pass Rate']
    for i, h in enumerate(headers, 1):
        ws_sum.cell(row=6, column=i, value=h)
    style_header(ws_sum, 6, 6)

    # 유형별 집계 (동적)
    type_groups = {}
    for r in results:
        key = r['type']
        if '25년' in r.get('folder', ''):
            key += ' 25년'
        elif '26년' in r.get('folder', ''):
            key += ' 26년'
        type_groups.setdefault(key, []).append(r)

    row = 7
    for cat, items in sorted(type_groups.items()):
        total = len(items)
        p = sum(1 for r in items if r['overall'] == 'Pass')
        e = sum(1 for r in items if r['overall'] == 'Exception')
        f = sum(1 for r in items if r['overall'] == 'Fail')
        ws_sum.cell(row=row, column=1, value=cat)
        ws_sum.cell(row=row, column=2, value=total)
        ws_sum.cell(row=row, column=3, value=p)
        ws_sum.cell(row=row, column=4, value=e)
        ws_sum.cell(row=row, column=5, value=f)
        ws_sum.cell(row=row, column=6, value=f'{p / total * 100:.1f}%' if total > 0 else 'N/A')
        style_data(ws_sum, row, 6)
        for c in range(2, 7):
            ws_sum.cell(row=row, column=c).alignment = center
        row += 1

    # 합계
    ws_sum.cell(row=row, column=1, value='합계')
    ws_sum.cell(row=row, column=1).font = data_font_bold
    ws_sum.cell(row=row, column=2, value=len(results))
    tp = sum(1 for r in results if r['overall'] == 'Pass')
    te = sum(1 for r in results if r['overall'] == 'Exception')
    tf = sum(1 for r in results if r['overall'] == 'Fail')
    ws_sum.cell(row=row, column=3, value=tp)
    ws_sum.cell(row=row, column=4, value=te)
    ws_sum.cell(row=row, column=5, value=tf)
    ws_sum.cell(row=row, column=6, value=f'{tp / len(results) * 100:.1f}%')
    style_data(ws_sum, row, 6, data_font_bold)
    for c in range(2, 7):
        ws_sum.cell(row=row, column=c).alignment = center

    # Exception/Fail 목록
    row += 2
    ws_sum.cell(row=row, column=1, value='2. Exception / Fail 목록')
    ws_sum.cell(row=row, column=1).font = subtitle_font
    row += 1
    exc_headers = ['No.', '폴더명', '유형', '종합결과', 'Exception/Fail 사유', '관련 테스트']
    for i, h in enumerate(exc_headers, 1):
        ws_sum.cell(row=row, column=i, value=h)
    style_header(ws_sum, row, 6)

    row += 1
    exc_no = 1
    for r in results:
        if r['overall'] in ['Exception', 'Fail']:
            ws_sum.cell(row=row, column=1, value=exc_no)
            ws_sum.cell(row=row, column=2, value=r['folder'])
            ws_sum.cell(row=row, column=3, value=r['type'])
            ws_sum.cell(row=row, column=4, value=r['overall'])
            color_result(ws_sum.cell(row=row, column=4), r['overall'])
            ws_sum.cell(row=row, column=5, value='\n'.join(r.get('exceptions', [])))
            style_data(ws_sum, row, 6)
            row += 1
            exc_no += 1

    ws_sum.column_dimensions['A'].width = 22
    ws_sum.column_dimensions['B'].width = 14
    ws_sum.column_dimensions['C'].width = 10
    ws_sum.column_dimensions['D'].width = 12
    ws_sum.column_dimensions['E'].width = 50
    ws_sum.column_dimensions['F'].width = 25

    # ============ Sheet 2: 해외매출 Detail ============
    overseas = [r for r in results if r['type'] != '국내매출']
    if overseas:
        ws_o = wb.create_sheet('해외매출 상세')
        ws_o['A1'] = f'해외매출 TOD 상세 워크페이퍼 - {company_name}'
        ws_o['A1'].font = title_font

        o_headers = [
            'No.', '폴더명', '매출유형', 'CI No.', '거래처', '운임조건', '통화',
            'FC금액', 'KRW금액',
            'T1:PO\n결과', 'T1:PO 상세',
            'T2:Invoice\n결과', 'T2:Invoice 상세',
            'T3:선적\n결과', 'T3:선적 상세',
            'T4:Incoterms\n결과', 'T4:Incoterms 상세',
            'T5:POD\n결과', 'T5:POD 상세',
            'T6:금액\n결과', 'T6:금액 상세',
            '종합결과', 'Exception 사유', '증빙파일 목록'
        ]

        for i, h in enumerate(o_headers, 1):
            ws_o.cell(row=3, column=i, value=h)
        style_header(ws_o, 3, len(o_headers))

        row = 4
        for no, r in enumerate(overseas, 1):
            ws_o.cell(row=row, column=1, value=no)
            ws_o.cell(row=row, column=2, value=r['folder'])
            ws_o.cell(row=row, column=3, value=r['type'])
            ws_o.cell(row=row, column=4, value=r.get('ci_no_excel', ''))
            ws_o.cell(row=row, column=5, value=r.get('customer_excel', ''))
            ws_o.cell(row=row, column=6, value=r.get('incoterms_excel', ''))
            ws_o.cell(row=row, column=7, value=r.get('currency_excel', ''))
            fc = r.get('fc_amount_excel')
            ws_o.cell(row=row, column=8, value=float(fc) if fc else '')
            krw = r.get('krw_amount_excel')
            ws_o.cell(row=row, column=9, value=float(krw) if krw else '')

            tests = [
                ('test_1_po_order', 10, 11),
                ('test_2_invoice', 12, 13),
                ('test_3_shipping', 14, 15),
                ('test_4_incoterms_revenue', 16, 17),
                ('test_5_pod_arrival', 18, 19),
                ('test_6_amount_match', 20, 21),
            ]
            for tkey, col_r, col_d in tests:
                t = r.get(tkey, {})
                ws_o.cell(row=row, column=col_r, value=t.get('result', ''))
                ws_o.cell(row=row, column=col_d, value=t.get('detail', ''))
                color_result(ws_o.cell(row=row, column=col_r), t.get('result', ''))

            ws_o.cell(row=row, column=22, value=r['overall'])
            color_result(ws_o.cell(row=row, column=22), r['overall'])
            ws_o.cell(row=row, column=23, value='\n'.join(r.get('exceptions', [])))

            all_files = []
            for doc_type, docs in r.get('docs', {}).items():
                for d in docs:
                    all_files.append(f"[{doc_type}] {d['filename']}")
            ws_o.cell(row=row, column=24, value='\n'.join(all_files))

            style_data(ws_o, row, len(o_headers))
            for col in [1, 3, 6, 7, 10, 12, 14, 16, 18, 20, 22]:
                ws_o.cell(row=row, column=col).alignment = wrap_center
            row += 1

        widths = [5, 24, 8, 24, 20, 7, 5, 14, 16,
                  8, 50, 8, 50, 8, 50, 8, 55, 8, 45, 8, 55, 8, 35, 55]
        for i, w in enumerate(widths, 1):
            if i <= len(o_headers):
                ws_o.column_dimensions[get_column_letter(i)].width = w

    # ============ Sheet 3: 국내매출 Detail ============
    domestic = [r for r in results if r['type'] == '국내매출']
    if domestic:
        ws_d = wb.create_sheet('국내매출 상세')
        ws_d['A1'] = f'국내매출 TOD 상세 워크페이퍼 - {company_name}'
        ws_d['A1'].font = title_font

        d_headers = [
            'No.', '폴더명', '거래처', '공급가액(KRW)',
            '세금계산서\n결과', '세금계산서 상세',
            '거래명세서\n결과', '거래명세서 상세',
            '금액대조\n결과', '금액대조 상세',
            '종합결과', 'Exception 사유', '증빙파일 목록'
        ]

        for i, h in enumerate(d_headers, 1):
            ws_d.cell(row=3, column=i, value=h)
        style_header(ws_d, 3, len(d_headers))

        row = 4
        for no, r in enumerate(domestic, 1):
            ws_d.cell(row=row, column=1, value=no)
            ws_d.cell(row=row, column=2, value=r['folder'])
            ws_d.cell(row=row, column=3, value=r.get('customer_excel', ''))
            krw = r.get('krw_amount_excel')
            ws_d.cell(row=row, column=4, value=float(krw) if krw else '')

            for tkey, col_r, col_d in [('test_tax_invoice', 5, 6), ('test_transaction_stmt', 7, 8), ('test_amount_match', 9, 10)]:
                t = r.get(tkey, {})
                ws_d.cell(row=row, column=col_r, value=t.get('result', ''))
                ws_d.cell(row=row, column=col_d, value=t.get('detail', ''))
                color_result(ws_d.cell(row=row, column=col_r), t.get('result', ''))

            ws_d.cell(row=row, column=11, value=r['overall'])
            color_result(ws_d.cell(row=row, column=11), r['overall'])
            ws_d.cell(row=row, column=12, value='\n'.join(r.get('exceptions', [])))

            all_files = []
            for doc_type, docs in r.get('docs', {}).items():
                for d in docs:
                    all_files.append(f"[{doc_type}] {d['filename']}")
            ws_d.cell(row=row, column=13, value='\n'.join(all_files))

            style_data(ws_d, row, len(d_headers))
            for col in [1, 5, 7, 9, 11]:
                ws_d.cell(row=row, column=col).alignment = wrap_center
            row += 1

        d_widths = [5, 24, 20, 16, 10, 55, 10, 45, 10, 55, 10, 35, 55]
        for i, w in enumerate(d_widths, 1):
            ws_d.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"워크페이퍼 저장: {output_path}")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("사용법: python3 tod_workpaper.py <results.json> <output.xlsx> [회사명] [감사기간]")
        sys.exit(1)

    json_path = sys.argv[1]
    output_path = sys.argv[2]
    company = sys.argv[3] if len(sys.argv) > 3 else ""
    period = sys.argv[4] if len(sys.argv) > 4 else ""

    with open(json_path, 'r', encoding='utf-8') as f:
        results = json.load(f)

    generate_workpaper(results, output_path, company, period)
