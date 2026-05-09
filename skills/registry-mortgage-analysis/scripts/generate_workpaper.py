#!/usr/bin/env python3
"""부동산등기부등본 검토 워크페이퍼 생성 (4시트 엑셀)

사용법:
    python generate_workpaper.py <properties_json> <mortgages_json> <output_xlsx> \
        --company "피감사회사명" --period "감사기간" [--company-keywords kw1 kw2]

입력:
    properties_json — parse_registry_pdfs.py가 생성한 부동산 기본정보
    mortgages_json — parse_registry_pdfs.py가 생성한 근저당 상세

출력:
    4시트 구성 엑셀 워크페이퍼
"""

import json, re, os, sys, argparse
from collections import defaultdict
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system('pip install openpyxl --break-system-packages -q')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ===== 스타일 정의 =====
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill('solid', fgColor='2F5496')
hdr_font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=9)
dfont = Font(name='맑은 고딕', size=9)
bold9 = Font(name='맑은 고딕', bold=True, size=9)
bold10 = Font(name='맑은 고딕', bold=True, size=10)
norm10 = Font(name='맑은 고딕', size=10)
amt_fmt = '#,##0'

# 소유 구분 색상
company_fill = PatternFill('solid', fgColor='D6E4F0')       # 피감사회사 소유 (파란)
company_joint_fill = PatternFill('solid', fgColor='B4C7E7')  # 피감사회사 공동소유 (진파란)
other_fill = PatternFill('solid', fgColor='FFF2CC')           # 제3자 (노란)
sum_fill = PatternFill('solid', fgColor='E2EFDA')             # 요약 (녹색)
group_fill = PatternFill('solid', fgColor='FCE4D6')           # 공동담보 (주황)
total_fill = PatternFill('solid', fgColor='D9E2F3')           # 합계 (하늘)


def classify_ownership(prop, company_keywords):
    """소유 구분 판정"""
    owners = prop.get('소유자목록', [])
    is_company = prop.get('피감사회사소유', False)

    if not is_company and company_keywords:
        for o in owners:
            for kw in company_keywords:
                if kw in o.get('이름', ''):
                    is_company = True
                    break

    if not owners or len(owners) == 1:
        own_type = '단독'
    else:
        own_type = '공동'

    if own_type == '공동' and is_company:
        fill = company_joint_fill
        label = '공동'
    elif is_company:
        fill = company_fill
        label = 'Y'
    else:
        fill = other_fill
        label = 'N'

    # 지분
    if own_type == '공동' and owners:
        company_owners = [o for o in owners if any(kw in o.get('이름', '') for kw in (company_keywords or []))]
        share = company_owners[0]['지분'] if company_owners else ''
    elif is_company:
        share = '단독'
    else:
        share = '-'

    return is_company, own_type, label, share, fill


def build_joint_groups(mortgages):
    """공동담보 그룹 구성 (하이브리드: 공동담보목록 우선, 접수번호 폴백)"""
    groups = defaultdict(lambda: {'amt': 0, 'items': [], 'key_type': ''})
    for m in mortgages:
        key = m.get('공동담보목록', '')
        if key:
            kt = '목록'
        else:
            key = 'receipt_' + m.get('접수번호', 'unknown')
            kt = '접수' if m.get('공동담보', '') else '단독'
        if groups[key]['amt'] == 0:
            groups[key]['amt'] = m['채권최고액']
            groups[key]['key_type'] = kt
        groups[key]['items'].append(m)
    return groups


def create_sheet1(wb, mortgages, properties, company_keywords, company_name, period):
    """Sheet 1: 근저당 상세"""
    ws = wb.active
    ws.title = "1_근저당상세"

    prop_by_uid = {p['고유번호']: p for p in properties}

    ws.merge_cells('A1:Q1')
    ws['A1'] = '부동산등기부등본 근저당 상세 검토표'
    ws['A1'].font = Font(name='맑은 고딕', bold=True, size=12)
    ws.merge_cells('A2:Q2')
    ws['A2'] = f'피감사회사: {company_name} | 감사기간: {period} | 열람일: {date.today().strftime("%Y.%m.%d")} | 현행 근저당 (하단 요약 섹션 기준)'
    ws['A2'].font = Font(name='맑은 고딕', size=8, color='666666')

    headers = ['No', '고유번호', '유형', '소재지', '소유자', '회사소유', '지분',
               '순위번호', '설정일자', '접수정보', '채권최고액', '근저당권자', '채무자',
               '공동담보목록', '공동담보구분', '공동담보', '비고']
    widths = [5, 18, 5, 32, 18, 7, 14, 7, 12, 22, 18, 18, 18, 22, 8, 22, 12]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    # 부동산별 그룹핑
    mort_by_uid = defaultdict(list)
    for m in mortgages:
        mort_by_uid[m['고유번호']].append(m)

    row = 4
    for prop in sorted(properties, key=lambda x: x['No']):
        uid = prop['고유번호']
        morts = mort_by_uid.get(uid, [])
        is_co, own_type, co_label, share, rfill = classify_ownership(prop, company_keywords)
        owner_disp = prop.get('현재소유자', '')

        if not morts:
            vals = [prop['No'], uid, prop.get('유형', '-'), prop.get('소재지', '-'),
                    owner_disp or '-', co_label, share,
                    '-', '-', '-', 0, '-', '-', '-', '-', '-', '근저당 없음']
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.font = dfont; c.fill = rfill; c.border = border
                if col == 11: c.number_format = amt_fmt; c.alignment = Alignment(horizontal='right')
                elif col in (1, 3, 6, 8): c.alignment = Alignment(horizontal='center')
            row += 1
        else:
            for m in morts:
                jt_list = m.get('공동담보목록', '')
                jt_raw = m.get('공동담보', '')
                if jt_list:
                    jt_type = '목록'
                elif jt_raw:
                    jt_type = '접수'
                else:
                    jt_type = '단독'

                vals = [
                    prop['No'], uid,
                    prop.get('유형', '') or '-', prop.get('소재지', '') or '-',
                    owner_disp or '-', co_label, share,
                    int(m['순위번호']),
                    m.get('설정일자', '') or '-',
                    m.get('접수정보', '') or '-',
                    m['채권최고액'],
                    m.get('근저당권자', '') or '-',
                    m.get('채무자', '') or '-',
                    jt_list or '-', jt_type, jt_raw or '-', '-',
                ]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.font = dfont; c.fill = rfill; c.border = border
                    if col == 11: c.number_format = amt_fmt; c.alignment = Alignment(horizontal='right')
                    elif col in (1, 3, 6, 8, 15): c.alignment = Alignment(horizontal='center')
                row += 1

    # 합계
    ws.cell(row=row, column=10, value='합계').font = bold9
    ws.cell(row=row, column=10).border = border
    ws.cell(row=row, column=10).alignment = Alignment(horizontal='center')
    c = ws.cell(row=row, column=11)
    c.value = f'=SUM(K4:K{row - 1})'
    c.font = bold9; c.number_format = amt_fmt; c.border = border
    c.alignment = Alignment(horizontal='right')

    ws.auto_filter.ref = f'A3:Q{row - 1}'
    ws.freeze_panes = 'A4'
    return ws


def create_sheet2(wb, mortgages, properties, company_keywords, company_name, period):
    """Sheet 2: 부동산 종합표"""
    ws = wb.create_sheet("2_부동산종합표")

    mort_by_uid = defaultdict(list)
    for m in mortgages:
        mort_by_uid[m['고유번호']].append(m)

    ws.merge_cells('A1:R1')
    ws['A1'] = '부동산등기부등본 검토 종합표'
    ws['A1'].font = Font(name='맑은 고딕', bold=True, size=12)
    ws.merge_cells('A2:R2')
    ws['A2'] = f'피감사회사: {company_name} | 감사기간: {period} | 열람일: {date.today().strftime("%Y.%m.%d")}'
    ws['A2'].font = Font(name='맑은 고딕', size=8, color='666666')

    headers = ['No', '고유번호', '유형', '소재지', '소유자', '회사소유', '지분',
               '현행근저당(건)', '현행채권최고액합계', '근저당권자', '채무자', '공동담보',
               '현행가압류(건)', '가압류상세', '비고']
    widths = [5, 18, 5, 32, 18, 7, 14, 7, 18, 20, 18, 20, 7, 20, 12]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 4
    for prop in sorted(properties, key=lambda x: x['No']):
        uid = prop['고유번호']
        morts = mort_by_uid.get(uid, [])
        is_co, own_type, co_label, share, rfill = classify_ownership(prop, company_keywords)

        cred_list = ', '.join(sorted(set(m.get('근저당권자', '') for m in morts if m.get('근저당권자'))))
        debt_list = ', '.join(sorted(set(m.get('채무자', '') for m in morts if m.get('채무자'))))
        jt_list = ', '.join(sorted(set(m.get('공동담보목록', '') for m in morts if m.get('공동담보목록'))))

        vals = [
            prop['No'], uid, prop.get('유형', ''), prop.get('소재지', ''),
            prop.get('현재소유자', ''), co_label, share,
            len(morts), sum(m['채권최고액'] for m in morts),
            cred_list or '-', debt_list or '-', jt_list or '-',
            prop.get('현행_가압류_건수', 0), prop.get('갑구_특이사항', '-'), '-',
        ]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = dfont; c.fill = rfill; c.border = border
            if col == 9: c.number_format = amt_fmt; c.alignment = Alignment(horizontal='right')
            elif col in (1, 3, 6, 8, 13): c.alignment = Alignment(horizontal='center')
        row += 1

    # 합계
    ws.cell(row=row, column=7, value='합계').font = bold9
    ws.cell(row=row, column=7).border = border
    c = ws.cell(row=row, column=9)
    c.value = f'=SUM(I4:I{row - 1})'
    c.font = bold9; c.number_format = amt_fmt; c.border = border

    ws.auto_filter.ref = f'A3:O{row - 1}'
    ws.freeze_panes = 'A4'
    return ws


def create_sheet3(wb, mortgages, properties, jt_groups, company_name):
    """Sheet 3: 요약"""
    ws = wb.create_sheet("3_요약")
    ws['A1'] = '부동산등기부등본 검토 요약'
    ws['A1'].font = Font(name='맑은 고딕', bold=True, size=14)

    total_gross = sum(m['채권최고액'] for m in mortgages)
    total_net = sum(g['amt'] for g in jt_groups.values())

    co_count = sum(1 for p in properties if p.get('피감사회사소유'))
    seizure_count = sum(1 for p in properties if p.get('현행_가압류_건수', 0) > 0)

    r = 3
    items = [
        ('총 부동산 건수', len(properties)),
        (f'  {company_name} 소유', co_count),
        ('  제3자 소유', len(properties) - co_count),
        ('', ''),
        ('현행 근저당 건수', len(mortgages)),
        ('현행 채권최고액 합계 (Gross)', total_gross),
        ('', ''),
        ('순 담보금액 (공동담보 중복 제거)', ''),
        ('  총 채권최고액 (Gross)', total_gross),
        ('  공동담보 중복금액', total_gross - total_net),
        ('  순 담보금액 (Net)', total_net),
        ('  공동담보 그룹 수', len(jt_groups)),
        ('', ''),
        ('현행 가압류/압류 있는 부동산', f'{seizure_count}건' if seizure_count > 0 else '없음'),
    ]

    for label, val in items:
        ws.cell(row=r, column=1, value=label).font = Font(
            name='맑은 고딕', size=10,
            bold=bool(label and not label.startswith(' '))
        )
        c = ws.cell(row=r, column=2, value=val)
        c.font = norm10
        if isinstance(val, (int, float)) and val > 1000:
            c.number_format = '#,##0'
        r += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 55

    # 근저당권자별 집계
    r += 1
    ws.cell(row=r, column=1, value='근저당권자별 집계').font = Font(name='맑은 고딕', bold=True, size=11)
    r += 1
    for h, ci in [('근저당권자', 1), ('건수', 2), ('채권최고액합계', 3)]:
        ws.cell(row=r, column=ci, value=h).font = bold9
    ws.column_dimensions['C'].width = 20
    r += 1

    cred_stats = defaultdict(lambda: {'건수': 0, '합계': 0})
    for m in mortgages:
        key = m.get('근저당권자', '') or '(미파싱)'
        cred_stats[key]['건수'] += 1
        cred_stats[key]['합계'] += m['채권최고액']
    for cred, s in sorted(cred_stats.items(), key=lambda x: -x[1]['합계']):
        ws.cell(row=r, column=1, value=cred).font = dfont
        ws.cell(row=r, column=2, value=s['건수']).font = dfont
        c = ws.cell(row=r, column=3, value=s['합계'])
        c.font = dfont; c.number_format = '#,##0'
        r += 1

    return ws


def create_sheet4(wb, mortgages, properties, jt_groups, company_name, period):
    """Sheet 4: 담보금액 분석"""
    ws = wb.create_sheet("4_담보금액분석")

    total_gross = sum(m['채권최고액'] for m in mortgages)
    total_net = sum(g['amt'] for g in jt_groups.values())

    ws.merge_cells('A1:L1')
    ws['A1'] = '공동담보 고려 순 담보금액 분석'
    ws['A1'].font = Font(name='맑은 고딕', bold=True, size=12)
    ws.merge_cells('A2:L2')
    ws['A2'] = f'피감사회사: {company_name} | 감사기간: {period} | 산출방식: 공동담보목록 번호 우선, 미파싱시 접수번호 기준'
    ws['A2'].font = Font(name='맑은 고딕', size=8, color='666666')

    # Section A: 요약
    r = 4
    for label, val in [('총 채권최고액 (Gross)', total_gross),
                        ('공동담보 중복금액', total_gross - total_net),
                        ('순 담보금액 (Net)', total_net)]:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = bold10; c1.border = border; c1.fill = sum_fill
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = bold10 if '순' in label else norm10
        c2.number_format = '#,##0'; c2.border = border; c2.fill = sum_fill
        c2.alignment = Alignment(horizontal='right')
        r += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22

    # Section B: 그룹별 상세
    r += 1
    ws.merge_cells(f'A{r}:L{r}')
    ws.cell(row=r, column=1, value='공동담보 그룹별 순 담보금액 산출 내역').font = Font(name='맑은 고딕', bold=True, size=11)
    r += 1

    h4 = ['No', '공동담보 그룹', '식별방식', '근저당권자', '채무자',
          '채권최고액(건당)', '대상물건수', '대상물건(No)', 'Gross합계', '순담보금액', '중복제거액', '비고']
    w4 = [4, 28, 8, 18, 18, 18, 8, 25, 18, 18, 18, 15]

    for col, (h, w) in enumerate(zip(h4, w4), 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    hdr_row = r
    r += 1
    data_start = r

    group_no = 0
    for key, g in sorted(jt_groups.items(), key=lambda x: -x[1]['amt']):
        group_no += 1
        items = g['items']
        creds = ', '.join(sorted(set(m.get('근저당권자', '') for m in items if m.get('근저당권자'))))
        debts = ', '.join(sorted(set(m.get('채무자', '') for m in items if m.get('채무자'))))
        nos = sorted(set(m['No'] for m in items))
        nos_str = ', '.join(str(n) for n in nos)
        gross = sum(m['채권최고액'] for m in items)
        net = g['amt']
        dedup = gross - net

        row_fill = group_fill if len(items) > 1 else PatternFill('solid', fgColor='FFFFFF')
        remark = f'{len(items)}건 공동담보 → 1건 인정' if len(items) > 1 else '단독담보'

        vals = [group_no, key, g['key_type'], creds, debts,
                net, len(items), nos_str, gross, net, dedup, remark]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = dfont; c.fill = row_fill; c.border = border
            if col in (6, 9, 10, 11):
                c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
            elif col in (1, 3, 7):
                c.alignment = Alignment(horizontal='center')
        r += 1

    # 합계
    ws.cell(row=r, column=2, value='합계').font = bold9
    ws.cell(row=r, column=2).border = border; ws.cell(row=r, column=2).fill = total_fill
    ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=7, value=len(mortgages)).font = bold9
    ws.cell(row=r, column=7).border = border; ws.cell(row=r, column=7).fill = total_fill
    ws.cell(row=r, column=7).alignment = Alignment(horizontal='center')

    for ci, cl in [(9, 'I'), (10, 'J'), (11, 'K')]:
        c = ws.cell(row=r, column=ci)
        c.value = f'=SUM({cl}{data_start}:{cl}{r - 1})'
        c.font = bold9; c.number_format = '#,##0'; c.border = border
        c.fill = total_fill; c.alignment = Alignment(horizontal='right')

    for ci in [1, 3, 4, 5, 6, 8, 12]:
        ws.cell(row=r, column=ci).border = border
        ws.cell(row=r, column=ci).fill = total_fill

    r += 2
    # Section C: 공동담보 그룹 상세 전개
    ws.merge_cells(f'A{r}:L{r}')
    ws.cell(row=r, column=1, value='공동담보 그룹별 대상 물건 상세').font = Font(name='맑은 고딕', bold=True, size=11)
    r += 1

    h4b = ['그룹No', '공동담보 그룹', '물건No', '고유번호', '유형', '소재지',
           '순위번호', '설정일자', '접수정보', '채권최고액', '근저당권자', '채무자']
    for col, h in enumerate(h4b, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    r += 1

    group_no = 0
    for key, g in sorted(jt_groups.items(), key=lambda x: -x[1]['amt']):
        group_no += 1
        if len(g['items']) < 2:
            continue
        for m in g['items']:
            vals = [group_no, key, m['No'], m['고유번호'],
                    m.get('유형', ''), m.get('소재지', ''),
                    int(m['순위번호']), m.get('설정일자', '-'),
                    m.get('접수정보', '-'), m['채권최고액'],
                    m.get('근저당권자', '-'), m.get('채무자', '-')]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.font = dfont; c.fill = group_fill; c.border = border
                if col == 10:
                    c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
                elif col in (1, 3, 5, 7):
                    c.alignment = Alignment(horizontal='center')
            r += 1

    ws.auto_filter.ref = f'A{hdr_row}:L{hdr_row + group_no}'
    ws.freeze_panes = f'A{hdr_row + 1}'
    return ws


def main():
    parser = argparse.ArgumentParser(description='부동산등기부등본 워크페이퍼 생성')
    parser.add_argument('properties', help='부동산 기본정보 JSON')
    parser.add_argument('mortgages', help='근저당 상세 JSON')
    parser.add_argument('output', help='출력 엑셀 파일 경로')
    parser.add_argument('--company', required=True, help='피감사회사명')
    parser.add_argument('--period', required=True, help='감사기간')
    parser.add_argument('--company-keywords', nargs='+', default=[], help='피감사회사 관련 키워드')
    args = parser.parse_args()

    with open(args.properties, 'r', encoding='utf-8') as f:
        properties = json.load(f)
    with open(args.mortgages, 'r', encoding='utf-8') as f:
        mortgages = json.load(f)

    print(f"부동산: {len(properties)}건, 근저당: {len(mortgages)}건")

    jt_groups = build_joint_groups(mortgages)
    total_gross = sum(m['채권최고액'] for m in mortgages)
    total_net = sum(g['amt'] for g in jt_groups.values())

    wb = Workbook()
    create_sheet1(wb, mortgages, properties, args.company_keywords, args.company, args.period)
    create_sheet2(wb, mortgages, properties, args.company_keywords, args.company, args.period)
    create_sheet3(wb, mortgages, properties, jt_groups, args.company)
    create_sheet4(wb, mortgages, properties, jt_groups, args.company, args.period)

    wb.save(args.output)
    print(f"\n워크페이퍼 저장: {args.output}")
    print(f"  Sheet1 근저당상세")
    print(f"  Sheet2 부동산종합표")
    print(f"  Sheet3 요약")
    print(f"  Sheet4 담보금액분석")
    print(f"\n총 채권최고액 (Gross): {total_gross:>20,}원")
    print(f"공동담보 중복:         {total_gross - total_net:>20,}원")
    print(f"순 담보금액 (Net):     {total_net:>20,}원")
    print(f"공동담보 그룹:         {len(jt_groups)}개")


if __name__ == '__main__':
    main()
