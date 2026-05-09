"""Self-test: 입력 DSD에 대해 convert → verify → 라운드트립 메타 sanity 체크.

회귀 테스트 용도 — sample 데이터 1개 받아 패키지가 install 없이도
정상 동작하는지 입증한다. sample 데이터가 없으면 안내하고 종료.

usage:
    python examples/test_self.py <any.dsd>
"""

from __future__ import annotations

import sys
import tempfile
import time
from pathlib import Path

# install 없이 단독 실행 가능하게 src/ 경로 추가
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from dsd_to_xlsx import convert, verify, __version__


def main() -> int:
    print(f"dsd-to-xlsx self-test  (version {__version__})")
    print()

    if len(sys.argv) < 2:
        # 자동으로 samples/ 안 첫 .dsd 찾기
        samples = Path(__file__).resolve().parent.parent / "samples"
        candidates = list(samples.rglob("*.dsd")) if samples.exists() else []
        if not candidates:
            print("usage: python examples/test_self.py <input.dsd>", file=sys.stderr)
            print("(또는 samples/ 폴더에 .dsd 1개 두고 인자 없이 실행)", file=sys.stderr)
            return 1
        src = candidates[0]
        print(f"[auto] sample 사용: {src.name}")
    else:
        src = Path(sys.argv[1])
        if not src.exists():
            print(f"입력 없음: {src}", file=sys.stderr)
            return 1

    with tempfile.TemporaryDirectory(prefix="dsd_self_test_") as td:
        dst = Path(td) / (src.stem + ".xlsx")

        # Phase A: convert
        t = time.time()
        out = convert(src, dst)
        t_convert = time.time() - t
        size_kb = out.stat().st_size / 1024
        print(f"[A] convert  : {t_convert:6.2f}s  ({size_kb:6.1f} KB)")

        # Phase B: verify
        t = time.time()
        report = verify(out)
        t_verify = time.time() - t
        print(f"[B] verify   : {t_verify:6.2f}s")

        print()
        print(report.format())
        print()

        # Phase C: 라운드트립 메타 sanity
        from openpyxl import load_workbook
        wb = load_workbook(out, data_only=True)
        ck = {
            "_STRUCTURE non-empty": wb["_STRUCTURE"].max_row > 0,
            "_META has DOCUMENT-HEADER.regname": any(
                isinstance(wb["_META"].cell(r, 1).value, str)
                and "regname" in str(wb["_META"].cell(r, 1).value).lower()
                for r in range(1, wb["_META"].max_row + 1)
            ),
            "_CELLMAP rows >= 100": wb["_CELLMAP"].max_row >= 100,
        }
        print("[C] 라운드트립 메타 sanity:")
        all_pass = True
        for name, ok in ck.items():
            mark = "OK" if ok else "FAIL"
            print(f"      [{mark}] {name}")
            if not ok:
                all_pass = False

        print()
        overall = report.ok and all_pass
        print(f"=== {'PASS' if overall else 'FAIL'} ===")
        return 0 if overall else 2


if __name__ == "__main__":
    sys.exit(main())
