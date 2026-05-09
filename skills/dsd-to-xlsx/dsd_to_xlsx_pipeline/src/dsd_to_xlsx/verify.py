"""변환 결과 xlsx 검증 — 시트 구조·메타·누락 체크.

dsd → xlsx 변환 직후 자동으로 호출되어 산출물의 완전성을 확인한다.
누락 항목은 경고로 표시되며 후속 작업(footing 검증 등) 전 사용자가 인지할 수 있다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_FS = ['재무상태표', '포괄손익계산서', '자본변동표', '현금흐름표']
REQUIRED_META = ['_STRUCTURE', '_META', '_EXTRACTIONS', '_ACODES', '_CELLMAP']
NOTE_RE = re.compile(r'^(\d+)\.\s+')
PERIOD_RE = re.compile(r'제\s*(\d+)\s*\((당|전)\)\s*기\s*[:\s]*(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일')
UNIT_KW_RE = re.compile(r'단\s*위[^가-힣]*([가-힣]+)')


@dataclass
class VerifyReport:
    xlsx_path: Path
    ok: bool = True
    total_sheets: int = 0
    total_cells: int = 0

    fs_present: list[str] = field(default_factory=list)
    missing_fs: list[str] = field(default_factory=list)
    missing_meta: list[str] = field(default_factory=list)
    note_count: int = 0
    note_numbers: list[int] = field(default_factory=list)
    missing_note_numbers: list[int] = field(default_factory=list)
    empty_sheets: list[str] = field(default_factory=list)

    company: str | None = None
    fiscal_year: int | None = None
    closing_date: str | None = None
    unit: str | None = None

    def has_issues(self) -> bool:
        return bool(
            self.missing_fs
            or self.missing_meta
            or self.missing_note_numbers
            or self.empty_sheets
        )

    def format(self) -> str:
        lines = []
        status = '[PASS]' if self.ok else '[ISSUES]'
        lines.append(f'=== Verify {status}: {self.xlsx_path.name} ===')
        lines.append(
            f'  Sheets: {self.total_sheets} '
            f'(FS: {len(self.fs_present)}/4, Notes: {self.note_count}, '
            f'Meta: {5 - len(self.missing_meta)}/5)'
        )
        lines.append(f'  Cells (rough): {self.total_cells:,}')
        lines.append(f'  Company: {self.company or "(missing)"}')
        period = ''
        if self.fiscal_year is not None:
            period = f'제{self.fiscal_year}기'
        if self.closing_date:
            period += f' / {self.closing_date}'
        lines.append(f'  Period:  {period or "(missing)"}')
        lines.append(f'  Unit:    {self.unit or "(missing)"}')

        if self.missing_fs:
            lines.append(f'  WARN: Missing FS sheets: {", ".join(self.missing_fs)}')
        if self.missing_meta:
            lines.append(f'  WARN: Missing meta sheets: {", ".join(self.missing_meta)}')
        if self.missing_note_numbers:
            lines.append(
                f'  WARN: Missing note numbers in 1..{max(self.note_numbers) if self.note_numbers else 0}: '
                f'{self.missing_note_numbers}'
            )
        if self.empty_sheets:
            lines.append(f'  WARN: Empty sheets: {", ".join(self.empty_sheets)}')
        if not self.has_issues():
            lines.append('  All structural checks passed.')
        return '\n'.join(lines)


def _extract_period_unit(ws) -> tuple[int | None, str | None, str | None]:
    """재무상태표 첫 15행에서 기수·결산일·단위 추출."""
    fy = closing = unit = None
    for r in range(1, min(16, (ws.max_row or 0) + 1)):
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            if fy is None:
                m = PERIOD_RE.search(v)
                if m and m.group(2) == '당':
                    fy = int(m.group(1))
                    closing = f'{m.group(3)}-{int(m.group(4)):02d}-{int(m.group(5)):02d}'
            if unit is None:
                m = UNIT_KW_RE.search(v)
                if m:
                    cand = m.group(1)
                    if cand in ('원', '천원', '백만원', '천만원'):
                        unit = cand
    return fy, closing, unit


def _extract_company_from_meta(wb) -> str | None:
    """_META 시트에서 DOCUMENT-HEADER.regname 추출."""
    if '_META' not in wb.sheetnames:
        return None
    ws = wb['_META']
    for r in range(1, (ws.max_row or 0) + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if isinstance(k, str) and 'regname' in k.lower():
            if v:
                return str(v)
    return None


def verify(xlsx_path) -> VerifyReport:
    """변환된 xlsx의 완전성을 검사한다.

    Args:
        xlsx_path: 검증 대상 xlsx 파일 경로

    Returns:
        VerifyReport — `.format()`으로 사람이 읽는 보고서, `.ok`로 통과 여부.
    """
    p = Path(xlsx_path)
    wb = load_workbook(p, data_only=True)
    names = wb.sheetnames
    rep = VerifyReport(xlsx_path=p, total_sheets=len(names))

    rep.fs_present = [s for s in REQUIRED_FS if s in names]
    rep.missing_fs = [s for s in REQUIRED_FS if s not in names]
    rep.missing_meta = [s for s in REQUIRED_META if s not in names]

    note_nums = sorted(
        int(NOTE_RE.match(n).group(1)) for n in names if NOTE_RE.match(n)
    )
    rep.note_numbers = note_nums
    rep.note_count = len(note_nums)
    if note_nums:
        expected = set(range(1, max(note_nums) + 1))
        rep.missing_note_numbers = sorted(expected - set(note_nums))

    cells = 0
    for n in names:
        ws = wb[n]
        r = ws.max_row or 0
        c = ws.max_column or 0
        cells += r * c
        if r == 0 or c == 0:
            rep.empty_sheets.append(n)
    rep.total_cells = cells

    rep.company = _extract_company_from_meta(wb)
    if '재무상태표' in names:
        fy, closing, unit = _extract_period_unit(wb['재무상태표'])
        rep.fiscal_year = fy
        rep.closing_date = closing
        rep.unit = unit

    rep.ok = not rep.has_issues()
    return rep
