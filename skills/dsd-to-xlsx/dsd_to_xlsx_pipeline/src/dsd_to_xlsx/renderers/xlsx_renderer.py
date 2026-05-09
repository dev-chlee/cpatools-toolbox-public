"""Render DsdDocument to XLSX with financial statements and notes on separate sheets."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from dsd_to_xlsx.extractors.number_parser import parse_number
from dsd_to_xlsx.extractors.table_extractor import find_tables
from dsd_to_xlsx.models import DocumentNode, DsdDocument, NoteSection
from dsd_to_xlsx.parsers.note_splitter import split_notes
from dsd_to_xlsx.renderers.xlsx_metadata import (
    AcodeEntry,
    CellMapEntry,
    write_acodes_sheet,
    write_cellmap_sheet,
    write_extractions_sheet,
    write_meta_sheet,
    write_structure_sheet,
)

# Financial statement sheet names by type keyword
FINANCIAL_SHEET_NAMES: dict[str, str] = {
    "재무상태표": "재무상태표",
    "포괄손익계산서": "포괄손익계산서",
    "자본변동표": "자본변동표",
    "현금흐름표": "현금흐름표",
}

_FONT_NAME = "맑은 고딕"
_FONT_SIZE = 10

_THIN_SIDE = Side(style="thin")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)
_HEADER_FILL = PatternFill(fgColor="DCDCDC", fill_type="solid")


def _font(*, bold: bool = False, color: str | None = None) -> Font:
    """Create a Font with the standard name and size."""
    kwargs: dict = {"name": _FONT_NAME, "size": _FONT_SIZE}
    if bold:
        kwargs["bold"] = True
    if color:
        kwargs["color"] = color
    return Font(**kwargs)

_VALID_HORIZONTAL = {"left", "center", "right", "justify", "distributed", "fill", "general", "centerContinuous"}
_VALID_VERTICAL = {"top", "center", "bottom", "justify", "distributed"}
_VALIGN_MAP = {"middle": "center"}
_HALIGN_MAP = {"centre": "center"}

# Sub-section pattern: (1), (2), ... at the start of text
_SUBSECTION_RE = re.compile(r"^\(\d+\)\s")
# Maximum column width for auto-sizing (characters)
_MAX_COL_WIDTH = 50


def render_xlsx(
    doc: DsdDocument,
    output_path: str | Path,
    *,
    source_path: str | Path | None = None,
) -> Path:
    """Render a DsdDocument to an XLSX file.

    Creates visible sheets for each financial statement and each note
    section, plus hidden technical sheets that let `dsd create`
    reconstruct a DART-compatible DSD from the xlsx alone:

      - ``_STRUCTURE``   — raw contents.xml
      - ``_META``        — meta.xml attributes as key/value
      - ``_EXTRACTIONS`` — meta.xml EXTRACT/ITEM rows
      - ``_ACODES``      — (sheet, cell) → ACODE mapping for editable cells

    Args:
        doc: Parsed DSD document.
        output_path: Where to write the xlsx.
        source_path: Original .dsd path — recorded in the 보고서정보
            sheet's '원본 DSD 파일명' field. When omitted, that field
            falls back to '(정보 없음)'.

    The hidden sheets are only populated when ``doc.raw_contents_xml`` /
    ``doc.raw_meta_xml`` are present (set by ``dsd_to_xlsx.parse()``).
    """
    output_path = Path(output_path)
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # Collect ACODE/AUNIT mappings (TE/TU only) and full cell↔XPath
    # mappings (TD/TH/TE/TU/P) so the template builder can patch the
    # XML when the user edits visible cells.
    acode_entries: list[AcodeEntry] = []
    cellmap_entries: list[CellMapEntry] = []

    # 0. 보고서정보 시트 — reader-facing summary of what this xlsx
    #    represents (company, period, statements present, notes count).
    #    Rendered first so it's the default tab when someone opens the file.
    _render_info_sheet(wb, doc, source_path=source_path)

    # 1. Financial statement sheets
    _render_financial_sheets(wb, doc, acode_entries, cellmap_entries)

    # 2. Note sheets
    notes = split_notes(doc.tree)
    for note in notes:
        _render_note_sheet(wb, note, acode_entries, cellmap_entries)

    # Ensure at least one sheet exists
    if len(wb.sheetnames) == 0:
        wb.create_sheet("Sheet1")

    # 3. Hidden metadata sheets (only if raw XML was preserved)
    if doc.raw_contents_xml:
        write_structure_sheet(wb, doc.raw_contents_xml)
    if doc.raw_meta_xml:
        write_meta_sheet(wb, doc.raw_meta_xml)
        write_extractions_sheet(wb, doc.raw_meta_xml)
    write_acodes_sheet(wb, acode_entries)
    write_cellmap_sheet(wb, cellmap_entries)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# 보고서정보 (report summary) sheet
# ---------------------------------------------------------------------------


_INFO_SHEET_NAME = "보고서정보"
_PLACEHOLDER = "(정보 없음)"


def _render_info_sheet(
    wb: Workbook, doc: DsdDocument, *, source_path: str | Path | None = None,
) -> None:
    """Insert the reader-facing summary sheet.

    Uses ``extract_report_metadata`` so all field parsing lives in one
    place. The sheet itself is a simple key/value table organized into
    four sections (identity, period, audit, content) separated by
    blank rows.
    """
    from datetime import datetime

    from dsd_to_xlsx.extractors.report_metadata import extract_report_metadata

    md = extract_report_metadata(doc, source_path=source_path)
    ws = wb.create_sheet(title=_INFO_SHEET_NAME)

    def _v(value) -> str:
        if value is None:
            return _PLACEHOLDER
        if isinstance(value, str):
            return value if value.strip() else _PLACEHOLDER
        return str(value)

    period_str = _PLACEHOLDER
    if md.current_period_start and md.current_period_end:
        period_str = f"{md.current_period_start} ~ {md.current_period_end}"
    elif md.current_period_end:
        period_str = md.current_period_end

    fs_marks = []
    for label, present in (
        ("재무상태표", md.has_balance_sheet),
        ("포괄손익계산서", md.has_income_statement),
        ("자본변동표", md.has_equity_changes),
        ("현금흐름표", md.has_cash_flow),
    ):
        fs_marks.append(f"{label} {'O' if present else 'X'}")
    fs_summary = (
        f"{len(md.statement_types_present)}종 — "
        + " / ".join(fs_marks)
    )

    sections: list[tuple[str, list[tuple[str, str]]]] = [
        ("식별", [
            ("회사명", _v(md.company_name)),
            ("회사코드", _v(md.company_code)),
            ("문서 종류", _v(md.document_name)),
            ("재무제표 구분", _v(md.consolidation_basis)),
        ]),
        ("회계기간 / 단위", [
            ("기수", _v(md.term_label)),
            ("회계기간 (당기)", period_str),
            ("단위", _v(md.unit)),
        ]),
        ("감사", [
            ("감사인", _v(md.auditor)),
            ("감사의견", _v(md.audit_opinion)),
        ]),
        ("내용 구성", [
            ("재무제표", fs_summary),
            ("주석 수", f"{md.notes_count}"),
            ("이미지 수", f"{md.images_count}"),
        ]),
        ("원본", [
            ("원본 DSD 파일명", _v(md.source_filename)),
            ("변환 시각", datetime.now().isoformat(timespec="seconds")),
        ]),
    ]

    row = 1
    title_cell = ws.cell(row=row, column=1, value="보고서 정보")
    title_cell.font = _font(bold=True)
    title_cell.font = Font(name=_FONT_NAME, size=14, bold=True)
    row += 2

    for section_label, items in sections:
        header = ws.cell(row=row, column=1, value=section_label)
        header.font = _font(bold=True)
        header.fill = _HEADER_FILL
        row += 1
        for key, value in items:
            k_cell = ws.cell(row=row, column=1, value=key)
            v_cell = ws.cell(row=row, column=2, value=value)
            k_cell.font = _font()
            v_cell.font = _font()
            k_cell.alignment = Alignment(horizontal="left", vertical="center")
            v_cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True,
            )
            row += 1
        row += 1  # blank spacer between sections

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    # Keep first visible row fixed when scrolling
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Financial statement rendering
# ---------------------------------------------------------------------------

def _render_financial_sheets(
    wb: Workbook, doc: DsdDocument,
    acode_entries: list[AcodeEntry] | None = None,
    cellmap_entries: list[CellMapEntry] | None = None,
) -> None:
    """Find financial statement table pairs and render each to a sheet."""
    tables = find_tables(doc.tree)

    i = 0
    while i < len(tables):
        table = tables[i]
        sheet_name = _identify_financial_table(table)
        if sheet_name is not None and i + 1 < len(tables):
            actual_name = sheet_name[:31]
            ws = wb.create_sheet(title=actual_name)
            # Render title table as a structured header (no cell mapping —
            # the visible structure differs from the original XML)
            row = _write_financial_header(ws, table)
            row += 1  # blank row between title and data
            data_start = row
            _write_table(
                ws, tables[i + 1], start_row=row,
                sheet_name=actual_name,
                acode_entries=acode_entries,
                cellmap_entries=cellmap_entries,
            )
            # Freeze panes below the data table header row
            data_header_end = _find_data_header_end(tables[i + 1], data_start)
            if data_header_end:
                ws.freeze_panes = f"A{data_header_end + 1}"
            i += 2
            continue
        i += 1


def _write_financial_header(ws: Worksheet, title_table: DocumentNode) -> int:
    """Render a financial statement title table as a structured header.

    No cell merging — each text goes to column 1 with appropriate alignment.
    """
    from dsd_to_xlsx.extractors.table_extractor import extract_table

    grid = extract_table(title_table)
    if not grid:
        return 1

    row = 1
    for grid_row in grid:
        # Collect non-empty cell texts
        texts = [t.strip() for t in grid_row if t.strip()]
        if not texts:
            continue

        if len(texts) == 1:
            cell = ws.cell(row=row, column=1, value=texts[0])
            cell.font = _font(bold=(row == 1))
            cell.alignment = Alignment(horizontal="center")
        elif len(texts) == 2:
            # Two texts (e.g., company name + unit) → col 1 left, col 2 right
            cell_left = ws.cell(row=row, column=1, value=texts[0])
            cell_left.font = _font()
            cell_left.alignment = Alignment(horizontal="left")
            cell_right = ws.cell(row=row, column=2, value=texts[1])
            cell_right.font = _font()
            cell_right.alignment = Alignment(horizontal="right")
        else:
            for ci, text in enumerate(texts):
                cell = ws.cell(row=row, column=ci + 1, value=text)
                cell.font = _font()

        row += 1

    return row


def _identify_financial_table(table: DocumentNode) -> str | None:
    """Check if a TABLE node is a financial statement title table.

    Uses the shared ``identify_statement_keyword`` helper so that the
    extractor, verifier, and xlsx renderer all agree on which tables
    are title tables — key for TOC (table-of-contents) tables that
    list multiple statement names; those must *not* be routed to a
    statement sheet.
    """
    from dsd_to_xlsx.extractors.financial_extractor import identify_statement_keyword
    from dsd_to_xlsx.extractors.table_extractor import extract_table

    grid = extract_table(table)
    keyword = identify_statement_keyword(grid)
    if keyword is None:
        return None
    return FINANCIAL_SHEET_NAMES[keyword]


def _find_data_header_end(table: DocumentNode, start_row: int) -> int | None:
    """Find the last THEAD row's Excel row number for freeze panes."""
    rows = _collect_table_rows(table)
    thead_indices = _get_thead_row_indices(table, rows)
    if not thead_indices:
        return None
    return start_row + max(thead_indices)


# ---------------------------------------------------------------------------
# Note sheet rendering
# ---------------------------------------------------------------------------

def _render_note_sheet(
    wb: Workbook, note: NoteSection,
    acode_entries: list[AcodeEntry] | None = None,
    cellmap_entries: list[CellMapEntry] | None = None,
) -> None:
    """Render a single NoteSection to its own sheet.

    Sheet name mirrors the original DSD heading — "1. 일반사항" rather
    than "주석01" — so a reader scanning the xlsx tabs recognizes
    each note immediately. Formatting + Excel-constraint handling
    lives in :func:`renderers.sheet_naming.note_sheet_name`.
    """
    from dsd_to_xlsx.renderers.sheet_naming import note_sheet_name

    actual_title = note_sheet_name(note.number, note.title)
    ws = wb.create_sheet(title=actual_title)

    ctx = _BlankRowContext()
    ctx.acode_entries = acode_entries
    ctx.cellmap_entries = cellmap_entries
    ctx.sheet_name = actual_title
    for node in note.nodes:
        ctx.row = _render_node_to_sheet_tracked(ws, node, ctx)

    # Auto-size column A based on paragraph text (A열에만 단락 텍스트가 들어감)
    _auto_size_note_column_a(ws)


def _auto_size_note_column_a(ws: Worksheet) -> None:
    """Set column A width based on the longest paragraph text (capped).

    Tables already have their own auto-sizing applied during _write_table().
    This only handles the A-column paragraph content that isn't from tables.
    """
    max_len = 0
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val is None:
            continue
        text = str(val)
        # Consider only the first line for width (wrap_text handles the rest)
        first_line_len = len(text.split("\n")[0])
        if first_line_len > max_len:
            max_len = first_line_len

    if max_len == 0:
        return

    width = min(max_len * 1.1 + 2, _MAX_COL_WIDTH)
    current = ws.column_dimensions["A"].width
    if current is None or current < width:
        ws.column_dimensions["A"].width = width


class _BlankRowContext:
    """Tracks blank row state to prevent consecutive blanks in note sheets.

    Also carries the current sheet name and a shared list of ACODE entries
    so that ``_write_table`` can record (sheet, cell) → ACODE mappings
    across all note sheets into a single collection for the ``_ACODES``
    hidden sheet.
    """

    def __init__(self) -> None:
        self.row: int = 1
        self.last_was_blank: bool = False
        self.sheet_name: str = ""
        self.acode_entries: list[AcodeEntry] | None = None
        self.cellmap_entries: list[CellMapEntry] | None = None

    def add_blank(self) -> int:
        """Request a blank row. Returns the next row (may be same if suppressed)."""
        if self.row <= 1 or self.last_was_blank:
            return self.row  # suppress leading or consecutive blanks
        self.last_was_blank = True
        self.row += 1
        return self.row

    def mark_content(self) -> None:
        """Mark that content was written at current row."""
        self.last_was_blank = False


def _render_node_to_sheet_tracked(
    ws: Worksheet, node: DocumentNode, ctx: _BlankRowContext,
) -> int:
    """Render a DocumentNode to a worksheet with blank row tracking."""
    if node.tag == "TABLE":
        # Blank row before table for visual separation
        ctx.add_blank()
        ctx.row = _write_table(
            ws, node, start_row=ctx.row, force_border=False,
            sheet_name=ctx.sheet_name,
            acode_entries=ctx.acode_entries,
            cellmap_entries=ctx.cellmap_entries,
        )
        ctx.mark_content()
        ctx.add_blank()
        return ctx.row
    elif node.tag == "PGBRK":
        return ctx.row
    elif node.tag == "P":
        text = _collect_all_text(node)
        if not text.strip():
            ctx.add_blank()
            return ctx.row
        # Record P → cellmap mapping at the first row this P will occupy.
        # Round-trip semantics: if the user changes the visible text, the
        # template builder will replace the entire P element's text.
        first_cell_coord = f"A{ctx.row}"
        ctx.row = _write_paragraph_lines_tracked(ws, text, node, ctx)
        if (
            ctx.cellmap_entries is not None and ctx.sheet_name
            and node.source_element is not None
        ):
            try:
                xpath = node.source_element.getroottree().getpath(node.source_element)
            except Exception:
                xpath = ""
            if xpath:
                ctx.cellmap_entries.append(CellMapEntry(
                    sheet=ctx.sheet_name,
                    cell=first_cell_coord,
                    xpath=xpath,
                    node_type="P",
                    original_text=text,
                ))
        return ctx.row
    elif node.tag in ("SECTION-1", "SECTION-2", "BODY", "DOCUMENT"):
        for child in node.children:
            ctx.row = _render_node_to_sheet_tracked(ws, child, ctx)
        return ctx.row
    elif node.tag == "TITLE":
        text = _collect_all_text(node)
        if text.strip():
            cell = ws.cell(row=ctx.row, column=1, value=text.strip())
            cell.font = _font(bold=True)
            ctx.mark_content()
            ctx.row += 1
        return ctx.row
    else:
        text = _collect_all_text(node)
        if text.strip():
            cell = ws.cell(row=ctx.row, column=1, value=text.strip())
            _apply_style(cell, node)
            ctx.mark_content()
            ctx.row += 1
        return ctx.row


def _write_paragraph_lines_tracked(
    ws: Worksheet, text: str, node: DocumentNode, ctx: _BlankRowContext,
) -> int:
    """Write paragraph lines with blank row tracking."""
    lines = text.split("\n")

    for line in lines:
        stripped = line.strip()
        if not stripped:
            ctx.add_blank()
            continue

        # Insert blank row before sub-section markers like (1), (2)
        is_subsection = bool(_SUBSECTION_RE.match(stripped))
        if is_subsection:
            ctx.add_blank()

        cell = ws.cell(row=ctx.row, column=1, value=stripped)
        _apply_style(cell, node)

        if is_subsection:
            cell.font = _font(bold=True)

        ctx.mark_content()
        ctx.row += 1

    return ctx.row


def _render_node_to_sheet(ws: Worksheet, node: DocumentNode, row: int) -> int:
    """Render a DocumentNode to a worksheet, returning the next available row."""
    if node.tag == "TABLE":
        # Blank row before table for visual separation
        if row > 1:
            row += 1
        row = _write_table(ws, node, start_row=row, force_border=False)
        row += 1  # blank row after table
        return row
    elif node.tag == "PGBRK":
        return row
    elif node.tag == "P":
        text = _collect_all_text(node)
        if not text.strip():
            # Empty <P></P> = blank row (paragraph separator in original XML)
            # Skip if at the very top of a sheet
            if row > 1:
                row += 1
            return row
        row = _write_paragraph_lines(ws, row, text, node)
        return row
    elif node.tag in ("SECTION-1", "SECTION-2", "BODY", "DOCUMENT"):
        for child in node.children:
            row = _render_node_to_sheet(ws, child, row)
        return row
    elif node.tag == "TITLE":
        text = _collect_all_text(node)
        if text.strip():
            cell = ws.cell(row=row, column=1, value=text.strip())
            cell.font = _font(bold=True)
            row += 1
        return row
    else:
        text = _collect_all_text(node)
        if text.strip():
            cell = ws.cell(row=row, column=1, value=text.strip())
            _apply_style(cell, node)
            row += 1
        return row


def _write_paragraph_lines(ws: Worksheet, row: int, text: str, node: DocumentNode) -> int:
    """Write a paragraph to the worksheet, splitting on newlines.

    Mirrors the original XML structure:
    - Single &cr; (\\n) = line break within a paragraph -> separate row
    - Double &cr;&cr; (\\n\\n) = paragraph break -> blank row between content
    - (N) sub-section markers get bold styling
    """
    lines = text.split("\n")

    for line in lines:
        stripped = line.strip()
        if not stripped:
            if row > 1:
                row += 1
            continue

        is_subsection = bool(_SUBSECTION_RE.match(stripped))
        if is_subsection and row > 1:
            row += 1

        cell = ws.cell(row=row, column=1, value=stripped)
        _apply_style(cell, node)

        if is_subsection:
            cell.font = _font(bold=True)

        row += 1

    return row


# ---------------------------------------------------------------------------
# Table writing
# ---------------------------------------------------------------------------

def _write_table(
    ws: Worksheet, table: DocumentNode, start_row: int,
    *,
    force_border: bool = False,
    sheet_name: str = "",
    acode_entries: list[AcodeEntry] | None = None,
    cellmap_entries: list[CellMapEntry] | None = None,
) -> int:
    """Write a TABLE node to the worksheet starting at start_row.

    Args:
        force_border: If True, apply thin borders to all cells regardless
            of the BORDER attribute (used for note tables).
        sheet_name: Sheet name to record in acode_entries / cellmap_entries.
        acode_entries: Optional list to append AcodeEntry for each TE/TU cell.
        cellmap_entries: Optional list to append CellMapEntry for every
            data cell (TD/TH/TE/TU) — used by the template builder to
            apply user edits back into the original XML by XPath.

    Returns the next available row after the table.
    """
    # Set column widths from COL elements
    _apply_col_widths(ws, table)

    has_border = force_border or table.attributes.get("BORDER") == "1"

    rows = _collect_table_rows(table)
    if not rows:
        return start_row

    # Track which cells are already occupied by rowspan
    occupied: dict[tuple[int, int], bool] = {}
    current_row = start_row

    # Determine which rows are in THEAD
    thead_row_indices = _get_thead_row_indices(table, rows)
    # If no explicit THEAD, treat the first row as header
    if not thead_row_indices and rows:
        thead_row_indices = {0}

    # Track max content width per column for auto-sizing
    col_content_widths: dict[int, float] = {}

    for row_idx, tr_node in enumerate(rows):
        cells = [c for c in tr_node.children if c.tag in ("TD", "TH", "TU", "TE")]
        col_idx = 1  # openpyxl is 1-based

        for cell_node in cells:
            # Skip occupied cells from previous rowspan
            while (current_row, col_idx) in occupied:
                col_idx += 1

            colspan = _int_attr(cell_node, "COLSPAN", 1)
            rowspan = _int_attr(cell_node, "ROWSPAN", 1)
            text = _collect_all_text(cell_node).strip()

            # Write cell value (try number conversion)
            value = _convert_value(text)
            cell = ws.cell(row=current_row, column=col_idx, value=value)

            # Apply number format
            if isinstance(value, (int, float)):
                if value < 0:
                    cell.number_format = '#,##0;[Red](#,##0)'
                else:
                    cell.number_format = '#,##0'

            # Apply styles
            is_header = row_idx in thead_row_indices or cell_node.tag == "TH"
            _apply_cell_style(cell, cell_node, has_border, is_header=is_header)

            # Record ACODE / AUNIT mapping for template-mode rebuild
            if acode_entries is not None and sheet_name and cell_node.tag in ("TE", "TU"):
                acode = cell_node.attributes.get("ACODE", "")
                aunit = cell_node.attributes.get("AUNIT", "")
                if acode or aunit:
                    acode_entries.append(AcodeEntry(
                        sheet=sheet_name,
                        cell=cell.coordinate,
                        acode=acode,
                        aunit=aunit,
                        element_type=cell_node.tag,
                    ))

            # Record (sheet, cell) → XPath mapping for every data cell.
            # The template builder uses this to apply user edits back into
            # the original XML, regardless of whether the cell has an ACODE.
            if (
                cellmap_entries is not None and sheet_name
                and cell_node.source_element is not None
            ):
                try:
                    xpath = cell_node.source_element.getroottree().getpath(
                        cell_node.source_element
                    )
                except Exception:
                    xpath = ""
                if xpath:
                    cellmap_entries.append(CellMapEntry(
                        sheet=sheet_name,
                        cell=cell.coordinate,
                        xpath=xpath,
                        node_type=cell_node.tag,
                        original_text=text,
                    ))

            # Track content width for auto-sizing (only for single-column cells)
            if colspan == 1 and text:
                text_width = len(text) * 1.2 + 2  # rough char width estimate
                col_content_widths[col_idx] = max(
                    col_content_widths.get(col_idx, 0), text_width
                )

            # Handle spans
            #   * Header cells: merge to preserve multi-row/col hierarchy
            #   * Data + rowspan: replicate value to each spanned row
            #     (no merge — CLAUDE.md reserves merging for headers).
            #     Replication keeps category labels like "총포괄이익"
            #     (rowspan=3 in 자본변동표 source XML) visible on every
            #     sub-row instead of vanishing into blank cells, which
            #     otherwise lose information during Excel filter/sort
            #     and cause DSD→xlsx verification to report missing
            #     labels.
            #   * Data + colspan only: first-cell-only (preserved — see
            #     test_data_colspan_not_merged).
            if colspan > 1 or rowspan > 1:
                end_row = current_row + rowspan - 1
                end_col = col_idx + colspan - 1
                if is_header:
                    ws.merge_cells(
                        start_row=current_row, start_column=col_idx,
                        end_row=end_row, end_column=end_col,
                    )
                elif rowspan > 1:
                    for r in range(current_row + 1, end_row + 1):
                        rep = ws.cell(row=r, column=col_idx, value=value)
                        if isinstance(value, (int, float)):
                            rep.number_format = cell.number_format
                        _apply_cell_style(
                            rep, cell_node, has_border, is_header=False,
                        )
                for r in range(current_row, end_row + 1):
                    for c in range(col_idx, end_col + 1):
                        if r == current_row and c == col_idx:
                            continue
                        # Replicated rowspan cells: they're written, not
                        # "occupied"-as-blank — mark them occupied so the
                        # next sibling TD doesn't try to overwrite.
                        occupied[(r, c)] = True

            col_idx += colspan

        current_row += 1

    # Auto-size columns that don't have explicit widths set
    for ci, content_width in col_content_widths.items():
        letter = get_column_letter(ci)
        current = ws.column_dimensions[letter].width
        auto_width = min(content_width, _MAX_COL_WIDTH)
        if current is None or current < auto_width:
            ws.column_dimensions[letter].width = auto_width

    return current_row


def _collect_table_rows(table: DocumentNode) -> list[DocumentNode]:
    """Collect all TR nodes from a TABLE, including those inside THEAD/TBODY."""
    rows: list[DocumentNode] = []
    for child in table.children:
        if child.tag == "TR":
            rows.append(child)
        elif child.tag in ("THEAD", "TBODY", "TFOOT"):
            for sub in child.children:
                if sub.tag == "TR":
                    rows.append(sub)
    return rows


def _get_thead_row_indices(table: DocumentNode, all_rows: list[DocumentNode]) -> set[int]:
    """Get the indices (in all_rows) of rows that belong to THEAD."""
    thead_rows: set[int] = set()
    for child in table.children:
        if child.tag == "THEAD":
            for sub in child.children:
                if sub.tag == "TR":
                    try:
                        idx = all_rows.index(sub)
                        thead_rows.add(idx)
                    except ValueError:
                        pass
    return thead_rows


def _apply_col_widths(ws: Worksheet, table: DocumentNode) -> None:
    """Apply column widths from COL WIDTH attributes.

    Never shrinks a column below its current width so that note sheet
    defaults (set earlier) are preserved.
    """
    col_idx = 1
    for child in table.children:
        if child.tag == "COLGROUP":
            for col in child.children:
                if col.tag == "COL":
                    _set_col_width(ws, col_idx, col.attributes.get("WIDTH", ""))
                    col_idx += 1
        elif child.tag == "COL":
            _set_col_width(ws, col_idx, child.attributes.get("WIDTH", ""))
            col_idx += 1


def _set_col_width(ws: Worksheet, col_idx: int, width_str: str) -> None:
    """Set column width from a pixel string, capped at _MAX_COL_WIDTH."""
    if not width_str:
        return
    try:
        px = int(width_str)
    except ValueError:
        return
    new_width = min(max(px / 7, 3), _MAX_COL_WIDTH)
    letter = get_column_letter(col_idx)
    current = ws.column_dimensions[letter].width
    if current is not None and current > new_width:
        return
    ws.column_dimensions[letter].width = new_width


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _apply_cell_style(cell, node: DocumentNode, has_border: bool, is_header: bool) -> None:
    """Apply styling to a cell based on the DSD node attributes and style."""
    bold = False
    fill = None
    alignment_kwargs: dict = {}

    # Style from USERMARK (font size/name ignored — using fixed font)
    if node.style:
        if node.style.bold:
            bold = True
        if node.style.bg_color:
            bg_hex = node.style.bg_color.lstrip("#")
            fill = PatternFill(fgColor=bg_hex, fill_type="solid")
        if node.style.align:
            alignment_kwargs["horizontal"] = node.style.align.lower()

    # XML ALIGN / VALIGN attributes
    align_attr = node.attributes.get("ALIGN", "").lower()
    valign_attr = node.attributes.get("VALIGN", "").lower()
    if align_attr and "horizontal" not in alignment_kwargs:
        align_attr = _HALIGN_MAP.get(align_attr, align_attr)
        if align_attr in _VALID_HORIZONTAL:
            alignment_kwargs["horizontal"] = align_attr
    if valign_attr:
        valign_attr = _VALIGN_MAP.get(valign_attr, valign_attr)
        if valign_attr in _VALID_VERTICAL:
            alignment_kwargs["vertical"] = valign_attr

    # Header styling
    if is_header:
        bold = True
        if fill is None:
            fill = _HEADER_FILL

    # Apply
    cell.font = _font(bold=bold)
    if fill:
        cell.fill = fill
    if alignment_kwargs:
        alignment_kwargs.setdefault("wrap_text", True)
        alignment_kwargs.setdefault("vertical", "top")
        cell.alignment = Alignment(**alignment_kwargs)
    if has_border:
        cell.border = _THIN_BORDER


def _apply_style(cell, node: DocumentNode) -> None:
    """Apply style from a P or SPAN node to a cell.

    Sets wrap_text=True and vertical="top" so long paragraphs display
    from the top of the cell when Excel auto-wraps.
    """
    bold = False

    if node.style and node.style.bold:
        bold = True

    # Also check first child SPAN for bold
    if not bold:
        for child in node.children:
            if child.tag == "SPAN" and child.style and child.style.bold:
                bold = True
                break

    cell.font = _font(bold=bold)

    horizontal = None
    if node.style and node.style.align:
        horizontal = node.style.align.lower()

    cell.alignment = Alignment(
        horizontal=horizontal, vertical="top", wrap_text=True,
    )


# ---------------------------------------------------------------------------
# Value conversion
# ---------------------------------------------------------------------------

def _convert_value(text: str):
    """Try to convert text to a numeric value for Excel."""
    if not text or text == "-" or text == "–":
        return text

    num = parse_number(text)
    if num is not None:
        # Return int when possible for cleaner Excel display
        if num == int(num):
            return int(num)
        return num
    return text


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------

def _collect_all_text(node: DocumentNode) -> str:
    """Recursively collect all text content from a node."""
    parts: list[str] = []
    if node.text:
        parts.append(node.text)
    for child in node.children:
        parts.append(_collect_all_text(child))
        if child.tail:
            parts.append(child.tail)
    return "".join(parts)


def _int_attr(node: DocumentNode, attr: str, default: int = 1) -> int:
    """Get an integer attribute value with default."""
    val = node.attributes.get(attr, "")
    try:
        return int(val)
    except ValueError:
        return default


