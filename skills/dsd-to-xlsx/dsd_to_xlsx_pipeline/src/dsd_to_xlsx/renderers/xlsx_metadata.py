"""Hidden-sheet writers that embed DSD metadata into an xlsx workbook.

The hidden sheets let `dsd create` reconstruct a DART-compatible DSD
from the xlsx alone (template mode):

- ``_STRUCTURE``     — original contents.xml (full text in A1)
- ``_META``          — meta.xml key/value attributes
- ``_EXTRACTIONS``   — meta.xml EXTRACT/ITEM rows
- ``_ACODES``        — (sheet, cell) → ACODE/AUNIT mapping for editable cells

All sheets are marked ``state='hidden'`` so ordinary users don't see them.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable

from lxml import etree
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# Hidden sheet names (single source of truth)
SHEET_STRUCTURE = "_STRUCTURE"
SHEET_META = "_META"
SHEET_EXTRACTIONS = "_EXTRACTIONS"
SHEET_ACODES = "_ACODES"
SHEET_CELLMAP = "_CELLMAP"


@dataclass
class AcodeEntry:
    """A single (sheet, cell) → ACODE/AUNIT mapping entry."""
    sheet: str
    cell: str  # A1 notation
    acode: str = ""
    aunit: str = ""
    element_type: str = ""  # TE | TU | EXTRACTION | TD


@dataclass
class CellMapEntry:
    """A single (sheet, cell) → XML XPath mapping entry.

    Used for round-trip editing: when the user changes a value in the
    visible xlsx sheet, the template builder finds the matching XML
    element by XPath and replaces its text. ``original_text`` lets the
    builder skip cells that haven't been edited.
    """
    sheet: str
    cell: str  # A1 notation
    xpath: str  # absolute lxml XPath, e.g. /DOCUMENT/BODY/.../TD[3]
    node_type: str  # TD | TH | TE | TU | P
    original_text: str = ""


def hide_sheet(ws: Worksheet) -> None:
    """Mark a worksheet as hidden."""
    ws.sheet_state = "hidden"


def write_structure_sheet(wb: Workbook, contents_xml: bytes) -> None:
    """Store the raw contents.xml as a single cell on ``_STRUCTURE``.

    xlsx cells have a practical length limit (~32,767 chars in Excel).
    For DSD files where contents.xml exceeds that, we split across rows
    of column A. Readers concatenate rows to reconstruct the original.
    """
    ws = wb.create_sheet(SHEET_STRUCTURE)
    hide_sheet(ws)

    text = contents_xml.decode("utf-8")
    # Excel cell character limit
    chunk = 30_000
    for i in range(0, len(text), chunk):
        ws.cell(row=(i // chunk) + 1, column=1, value=text[i:i + chunk])


def read_structure_sheet(wb: Workbook) -> bytes | None:
    """Reverse of write_structure_sheet — returns original contents.xml bytes."""
    if SHEET_STRUCTURE not in wb.sheetnames:
        return None
    ws = wb[SHEET_STRUCTURE]
    parts: list[str] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        if row[0] is None:
            continue
        parts.append(row[0])
    if not parts:
        return None
    return "".join(parts).encode("utf-8")


def write_meta_sheet(wb: Workbook, raw_meta_xml: bytes) -> None:
    """Dump all meta.xml attributes as key/value pairs on ``_META``.

    Captures every attribute from GENERATOR, DOCUMENT-HEADER,
    DOCUMENT-INFO, FILE-INFO, and DOCUMENT-IMG-LIST so we can
    reconstruct meta.xml byte-for-byte in template mode.
    """
    ws = wb.create_sheet(SHEET_META)
    hide_sheet(ws)

    ws.append(["key", "value"])

    if not raw_meta_xml:
        return

    root = etree.fromstring(raw_meta_xml)

    # Element+attribute flattening, namespaced as element.attr
    for elem in root.iter():
        tag = _local_tag(elem.tag)
        if tag in ("METAINFO", "EXTRACT", "ITEM"):
            continue  # handled separately
        for attr_name, attr_value in elem.attrib.items():
            key = f"{tag}.{attr_name}"
            ws.append([key, attr_value])

    # Image filenames (DOCUMENT-IMG-LIST/IMG)
    for img in root.iterfind(".//IMG"):
        filename = img.get("filename", "")
        if filename:
            ws.append(["IMG.filename", filename])


def read_meta_sheet(wb: Workbook) -> dict[str, list[str]]:
    """Reverse of write_meta_sheet — returns dict of element.attr → [values].

    Values are lists because some keys (like IMG.filename) can repeat.
    """
    result: dict[str, list[str]] = {}
    if SHEET_META not in wb.sheetnames:
        return result
    ws = wb[SHEET_META]
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, value = (row[0], row[1]) if len(row) >= 2 else (None, None)
        if key is None:
            continue
        result.setdefault(str(key), []).append("" if value is None else str(value))
    return result


def write_extractions_sheet(wb: Workbook, raw_meta_xml: bytes) -> None:
    """Dump all EXTRACT/ITEM rows on ``_EXTRACTIONS``.

    Columns: extract_type | name | value | tbl-grp-aclass | item_type
    """
    ws = wb.create_sheet(SHEET_EXTRACTIONS)
    hide_sheet(ws)

    ws.append(["extract_type", "name", "value", "tbl-grp-aclass", "item_type"])

    if not raw_meta_xml:
        return

    root = etree.fromstring(raw_meta_xml)
    for extract in root.iterfind(".//EXTRACT"):
        etype = extract.get("type", "")
        for item in extract.iterfind("./ITEM"):
            ws.append([
                etype,
                item.get("name", ""),
                item.get("value", ""),
                item.get("tbl-grp-aclass", ""),
                item.get("type", ""),
            ])


def read_extractions_sheet(wb: Workbook) -> list[dict[str, str]]:
    """Reverse of write_extractions_sheet — returns list of extraction row dicts."""
    if SHEET_EXTRACTIONS not in wb.sheetnames:
        return []
    ws = wb[SHEET_EXTRACTIONS]
    rows: list[dict[str, str]] = []
    header: list[str] | None = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(v) if v is not None else "" for v in row]
            continue
        if all(v is None for v in row):
            continue
        entry = {
            header[i]: ("" if row[i] is None else str(row[i]))
            for i in range(min(len(header), len(row)))
        }
        rows.append(entry)
    return rows


def write_acodes_sheet(wb: Workbook, entries: Iterable[AcodeEntry]) -> None:
    """Store (sheet, cell) → ACODE/AUNIT mappings on ``_ACODES``.

    This is the critical link that lets template-mode build locate
    user-edited values in visible sheets and inject them into the
    matching XML element.
    """
    ws = wb.create_sheet(SHEET_ACODES)
    hide_sheet(ws)

    ws.append(["sheet", "cell", "acode", "aunit", "element_type"])
    for e in entries:
        ws.append([e.sheet, e.cell, e.acode, e.aunit, e.element_type])


def read_acodes_sheet(wb: Workbook) -> list[AcodeEntry]:
    """Reverse of write_acodes_sheet."""
    if SHEET_ACODES not in wb.sheetnames:
        return []
    ws = wb[SHEET_ACODES]
    entries: list[AcodeEntry] = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if all(v is None for v in row):
            continue
        vals = ["" if v is None else str(v) for v in row]
        while len(vals) < 5:
            vals.append("")
        entries.append(AcodeEntry(
            sheet=vals[0], cell=vals[1],
            acode=vals[2], aunit=vals[3], element_type=vals[4],
        ))
    return entries


def write_cellmap_sheet(wb: Workbook, entries: Iterable[CellMapEntry]) -> None:
    """Store (sheet, cell) → XPath mappings on ``_CELLMAP``.

    This sheet is the critical link for round-trip editing: each row
    tells the template builder which XML element corresponds to which
    visible-sheet cell.
    """
    ws = wb.create_sheet(SHEET_CELLMAP)
    hide_sheet(ws)

    ws.append(["sheet", "cell", "xpath", "node_type", "original_text"])
    for e in entries:
        ws.append([e.sheet, e.cell, e.xpath, e.node_type, e.original_text])


def read_cellmap_sheet(wb: Workbook) -> list[CellMapEntry]:
    """Reverse of write_cellmap_sheet."""
    if SHEET_CELLMAP not in wb.sheetnames:
        return []
    ws = wb[SHEET_CELLMAP]
    entries: list[CellMapEntry] = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if all(v is None for v in row):
            continue
        vals = ["" if v is None else str(v) for v in row]
        while len(vals) < 5:
            vals.append("")
        entries.append(CellMapEntry(
            sheet=vals[0], cell=vals[1], xpath=vals[2],
            node_type=vals[3], original_text=vals[4],
        ))
    return entries


def _local_tag(tag: str) -> str:
    """Strip XML namespace (if any) from a tag name."""
    if tag.startswith("{"):
        return tag.split("}", 1)[1]
    return tag
