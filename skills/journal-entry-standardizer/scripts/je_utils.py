"""
분개장 표준화 — 범용 유틸리티 함수 모음
========================================
이 모듈은 분개장 형식·스키마에 의존하지 않는 **순수 변환 함수**만 포함한다.
특정 프로젝트의 컬럼명·시트 구조·검증 로직은 여기에 넣지 않는다.
LLM이 SKILL.md의 워크플로우를 따라 작업별 코드를 직접 작성할 때,
이 유틸 함수들을 import하여 빌딩 블록으로 사용한다.

사용법:
    import sys
    sys.path.insert(0, '<스킬경로>/scripts')
    from je_utils import *
"""
import re
from datetime import datetime


# ============================================================
# 1. 숫자 변환
# ============================================================

def safe_num(val):
    """다양한 형식의 값을 숫자(float)로 변환.
    None / 빈값 / '-' → 0
    콤마 포함 문자열('491,694') → 491694.0
    괄호 음수('(1,000)') → -1000.0
    """
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    s = str(val).replace(',', '').replace(' ', '').strip()
    if s == '' or s == '-':
        return 0
    # 괄호 음수 처리: (1000) → -1000
    m = re.match(r'^\((.+)\)$', s)
    if m:
        s = '-' + m.group(1)
    try:
        return float(s)
    except ValueError:
        return 0


# ============================================================
# 2. 텍스트 변환 / NULL 처리
# ============================================================

def safe_text(val):
    """셀 값을 텍스트로 변환. 빈 값이면 'NULL' 반환.
    원본 데이터에서 비어있는 셀을 명시적으로 표시하기 위함."""
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if s == '':
        return 'NULL'
    return s


# ============================================================
# 3. 날짜 변환
# ============================================================

def parse_date(val):
    """다양한 날짜 형식을 datetime으로 변환. 실패 시 None.
    지원: YYYY-MM-DD, YYYYMMDD, YYYY.MM.DD, YYYY/MM/DD,
          YYYY-MM-DD HH:MM:SS, 숫자(Excel serial date)
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    # Excel serial date (숫자)
    if isinstance(val, (int, float)):
        try:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return base + timedelta(days=int(val))
        except (ValueError, OverflowError):
            return None
    s = str(val).strip()
    for fmt in ['%Y-%m-%d', '%Y%m%d', '%Y.%m.%d', '%Y/%m/%d',
                '%Y-%m-%d %H:%M:%S', '%Y%m%d%H%M%S',
                '%m/%d/%Y', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt)
        except (ValueError, TypeError):
            continue
    return None


# ============================================================
# 4. 날짜 파생 값
# ============================================================

def derive_year(dt):
    """datetime → 연도 (int). None이면 None."""
    return dt.year if dt else None

def derive_month(dt):
    """datetime → 월 (int). None이면 None."""
    return dt.month if dt else None

def derive_quarter(dt):
    """datetime → 분기 (1~4). None이면 None."""
    if dt is None:
        return None
    return (dt.month - 1) // 3 + 1


# ============================================================
# 5. 계정코드 처리
# ============================================================

def parse_account_code(text):
    """계정과목 텍스트에서 코드와 명칭 분리.
    '[1350000]부가세대급금' → ('1350000', '부가세대급금')
    '부가세대급금'          → (None, '부가세대급금')
    None                   → (None, None)
    """
    if not text:
        return None, None
    text = str(text).strip()
    m = re.match(r'\[(\d+)\](.*)', text)
    if m:
        return m.group(1), m.group(2).strip()
    return None, text


def normalize_code(code):
    """계정코드를 순수 숫자 문자열로 정규화.
    12345 → '12345', '012345' → '12345', 12345.0 → '12345'
    0 → '0', None → None, '' → None, 비숫자 → None
    """
    if code is None:
        return None
    s = str(code).strip()
    if s == '':
        return None
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        if s.isdigit():
            return s
        return None


# ============================================================
# 6. 헤더 분석 (범용)
# ============================================================

def analyze_header(file_path, sheet_name=None, max_preview_rows=5):
    """엑셀 파일의 헤더 구조를 자동 분석.

    Returns:
        {
            'file': str,
            'sheet': str,
            'header_row': int (0-based),
            'skip_rows': int (데이터 시작 행),
            'headers': list[str|None],
            'num_cols': int,
            'num_data_rows': int (추정),
            'preview_rows': list[list],
            'pattern': str ('single_row'|'multi_row'|'subtotal_prefix'),
            'has_gubun_col': bool (차대분리형 여부),
            'has_embedded_code': bool (코드내장형 여부),
        }
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    actual_sheet = sheet_name or wb.sheetnames[0]

    all_rows = []
    total_rows = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= max_preview_rows + 10:
            all_rows.append(list(row))
        total_rows += 1
    wb.close()

    if not all_rows:
        return {'file': file_path, 'sheet': actual_sheet, 'error': 'empty file'}

    # 헤더 패턴 감지
    header_row = 0
    pattern = 'single_row'

    # SUBTOTAL / 합계 선행 감지
    row0_str = ' '.join(str(v) for v in all_rows[0] if v)
    if 'SUBTOTAL' in row0_str.upper() or '합계' in row0_str:
        header_row = 1
        pattern = 'subtotal_prefix'

    # 다중행 헤더 감지 (None 비율 30% 이상이면 그룹 헤더)
    if len(all_rows) >= 2:
        row0_nones = sum(1 for v in all_rows[header_row] if v is None)
        if row0_nones > len(all_rows[header_row]) * 0.3:
            pattern = 'multi_row'

    headers = [str(v).strip() if v else None for v in all_rows[header_row]]

    # 차대분리형(구분 컬럼) 감지
    has_gubun = any(h and '구분' in h for h in headers)

    # 코드내장형 감지: [1234567]계정명
    has_embedded_code = False
    for row in all_rows[header_row + 1:header_row + 4]:
        for v in row:
            if v and re.match(r'\[\d+\]', str(v)):
                has_embedded_code = True
                break

    skip_rows = header_row + (2 if pattern == 'multi_row' else 1)

    return {
        'file': file_path,
        'sheet': actual_sheet,
        'header_row': header_row,
        'skip_rows': skip_rows,
        'headers': headers,
        'num_cols': len(headers),
        'num_data_rows': total_rows - skip_rows,
        'preview_rows': all_rows[skip_rows:skip_rows + max_preview_rows],
        'pattern': pattern,
        'has_gubun_col': has_gubun,
        'has_embedded_code': has_embedded_code,
    }


# ============================================================
# 7. 엑셀 출력 헬퍼 (범용)
# ============================================================

# 표준 서식 상수
FONT_NAME = '맑은 고딕'
FONT_SIZE = 10
AMT_FMT = '#,##0;[Red](#,##0);-'
DATE_FMT = 'YYYY-MM-DD'
HDR_BG_COLOR = '4472C4'

def make_header_cell(ws, value):
    """write_only 워크시트용 헤더 셀 생성"""
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill, Alignment

    cell = WriteOnlyCell(ws, value=value)
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=HDR_BG_COLOR)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    return cell


def make_data_cell(ws, value, is_amount=False, is_date=False):
    """write_only 워크시트용 데이터 셀 생성"""
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font

    cell = WriteOnlyCell(ws, value=value)
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
    if is_amount:
        cell.number_format = AMT_FMT
    elif is_date and isinstance(value, datetime):
        cell.number_format = DATE_FMT
    return cell
