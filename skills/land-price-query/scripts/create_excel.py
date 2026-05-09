#!/usr/bin/env python3
"""
개별공시지가 조회 결과를 감사 워크페이퍼(엑셀)로 생성한다.

사용법:
    이 스크립트를 직접 실행하거나, create_workpaper() 함수를 import하여 사용.

    직접 실행 시: stdin으로 JSON 데이터를 받는다.
    python scripts/create_excel.py < workpaper_data.json

    JSON 형식:
    {
        "output_path": "/path/to/output.xlsx",
        "img_dir": "/path/to/images",
        "company_name": "(주)삼성전자",
        "audit_date": "2025.12.31",
        "query_date": "2026.03.24",
        "auditor_name": "홍길동 (공인회계사)",
        "properties": [
            {
                "seq": 1,
                "sido": "경기도",
                "sigungu": "OO시 OO구",
                "dong": "OO동",
                "jibun": "X-X",
                "sheet_name": "OO동 X-X",
                "img_file": "01_OO동_X-X.png",
                "location_full": "경기도 OO시 OO구 OO동 X-X",
                "price_2025": "1,234,567 원/㎡",
                "total_records": 38,
                "data": [
                    [2025, "경기도 OO시 OO구 OO동 X-X", "X-X번지", 1234567, "01월01일", "20250430"],
                    ...
                ]
            }
        ],
        "notes": [
            "개별공시지가는 매년 1월 1일 기준으로 산정되며, 2025년분은 2025.04.30에 공시되었습니다.",
            "OO시는 2025년 행정구역 분구로 인해 'OO신구'로 조회하였습니다 (OO면 소재)."
        ]
    }
"""
import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter


# ── 스타일 정의 ──
font_title = Font(name='맑은 고딕', size=14, bold=True)
font_subtitle = Font(name='맑은 고딕', size=10, bold=True)
font_header = Font(name='맑은 고딕', size=10, bold=True)
font_data = Font(name='맑은 고딕', size=10)
font_note = Font(name='맑은 고딕', size=9, italic=True, color='666666')
font_blue = Font(name='맑은 고딕', size=10, color='0000FF')
font_label = Font(name='맑은 고딕', size=10, bold=True)

fill_header = PatternFill('solid', fgColor='D6E4F0')
fill_light = PatternFill('solid', fgColor='F5F8FC')
fill_yellow = PatternFill('solid', fgColor='FFFFDD')
fill_title = PatternFill('solid', fgColor='1A5276')

thin = Side(style='thin', color='333333')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center')


def set_cell(ws, row, col, value, font=font_data, fill=None, alignment=align_center, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill: c.fill = fill
    if alignment: c.alignment = alignment
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c


def create_workpaper(data):
    """감사 워크페이퍼를 생성한다."""
    output_path = data["output_path"]
    img_dir = data.get("img_dir", os.path.dirname(output_path))
    props = data["properties"]
    notes = data.get("notes", [])

    wb = Workbook()

    # ════════════════════════════════════════════
    # Sheet 1: 요약
    # ════════════════════════════════════════════
    ws = wb.active
    ws.title = '요약'
    ws.sheet_properties.tabColor = '1A5276'

    for col, w in [(1,3),(2,18),(3,20),(4,30),(5,16),(6,18),(7,15),(8,12),(9,12)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # 타이틀
    ws.merge_cells('B1:H1')
    set_cell(ws, 1, 2, '개별공시지가 조회 결과',
             Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF'),
             fill_title, Alignment(horizontal='center', vertical='center'))
    for c in range(2, 9):
        ws.cell(row=1, column=c).fill = fill_title
    ws.row_dimensions[1].height = 40

    # 감사 정보
    r = 3
    info_items = [
        ('감사대상회사', data.get('company_name', '')),
        ('감사기준일', data.get('audit_date', '')),
        ('조회수행일', data.get('query_date', '')),
        ('조회수행자', data.get('auditor_name', '')),
    ]
    for label, val in info_items:
        set_cell(ws, r, 2, label, font_label, fill_header, align_left, border_all)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        set_cell(ws, r, 3, val, font_data, None, align_left, border_all)
        ws.cell(row=r, column=4).border = border_all
        r += 1

    # ── Section 1: 조회 신청 내역 ──
    r += 1
    set_cell(ws, r, 2, '1. 조회 신청 내역',
             Font(name='맑은 고딕', size=12, bold=True, color='1A5276'))
    r += 1
    set_cell(ws, r, 2,
             f'감사대상회사가 보유한 토지 {len(props)}건에 대해 개별공시지가를 조회하였습니다.',
             font_data, alignment=align_left)
    r += 1

    for ci, h in enumerate(['순번', '소재지 (시/도)', '소재지 (시/군/구)', '소재지 (읍/면/동)', '지번']):
        set_cell(ws, r, 2+ci, h, font_header, fill_header, align_center, border_all)
    r += 1

    for p in props:
        for ci, val in enumerate([p['seq'], p['sido'], p['sigungu'], p['dong'], p['jibun']]):
            set_cell(ws, r, 2+ci, val, font_data, None, align_center, border_all)
        r += 1

    # ── Section 2: 조회처 ──
    r += 1
    set_cell(ws, r, 2, '2. 조회처',
             Font(name='맑은 고딕', size=12, bold=True, color='1A5276'))
    r += 1

    source_items = [
        ('조회처', '부동산공시가격 알리미 (국토교통부)'),
        ('URL', 'https://www.realtyprice.kr/notice/gsindividual/search.htm'),
        ('조회구분', '개별지 공시지가 열람 > 지번 검색'),
        ('가격기준년도', f'{data.get("audit_date", "2025")[:4]}년'),
    ]
    for label, val in source_items:
        set_cell(ws, r, 2, label, font_label, fill_header, align_left, border_all)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        f = font_blue if 'http' in val else font_data
        set_cell(ws, r, 3, val, f, None, align_left, border_all)
        for cc in range(4, 7):
            ws.cell(row=r, column=cc).border = border_all
        r += 1

    r += 1
    set_cell(ws, r, 2,
             '※ 부동산공시가격 알리미는 국토교통부에서 운영하는 공식 부동산 공시가격 열람 시스템입니다.',
             font_note, alignment=align_left)

    # ── Section 3: 스크린샷 안내 ──
    r += 2
    set_cell(ws, r, 2, '3. 조회 결과 스크린샷',
             Font(name='맑은 고딕', size=12, bold=True, color='1A5276'))
    r += 1
    set_cell(ws, r, 2,
             '각 물건별 조회결과 이미지는 개별 시트 및 폴더 내 PNG 파일로 첨부되어 있습니다.',
             font_data, alignment=align_left)

    # ── Section 4: 결과 요약 ──
    r += 2
    set_cell(ws, r, 2, '4. 조회 결과 요약',
             Font(name='맑은 고딕', size=12, bold=True, color='1A5276'))
    r += 1

    headers4 = ['순번', '소재지', '지번', '개별공시지가\n(원/㎡)', '기준일자', '공시일자', '과거 데이터\n(건수)']
    for ci, h in enumerate(headers4):
        set_cell(ws, r, 2+ci, h, font_header, fill_header, align_center, border_all)
    ws.row_dimensions[r].height = 35
    r += 1

    for ri, p in enumerate(props):
        price_val = p['data'][0][3] if p['data'] else 0
        base_dt = p['data'][0][4] if p['data'] else ''
        ann_dt = p['data'][0][5] if p['data'] else ''
        row_data = [p['seq'], p['location_full'], f"{p['jibun']}번지",
                    price_val, base_dt, ann_dt, f"{p['total_records']}개"]
        bg = fill_yellow if ri == 0 else (fill_light if ri % 2 == 0 else None)
        for ci, val in enumerate(row_data):
            al = align_right if ci == 3 else align_center
            nf = '#,##0' if ci == 3 else None
            set_cell(ws, r, 2+ci, val,
                     Font(name='맑은 고딕', size=10, bold=(ci==3)), bg, al, border_all, nf)
        r += 1

    # 비고
    for note in notes:
        r += 1
        set_cell(ws, r, 2, f'※ {note}', font_note, alignment=align_left)

    # ════════════════════════════════════════════
    # 물건별 시트
    # ════════════════════════════════════════════
    for p in props:
        ws = wb.create_sheet(p["sheet_name"])
        ws.sheet_properties.tabColor = '4472C4'

        for col, w in [(1,3),(2,14),(3,35),(4,14),(5,18),(6,12),(7,12)]:
            ws.column_dimensions[get_column_letter(col)].width = w

        # 타이틀
        ws.merge_cells('B1:G1')
        set_cell(ws, 1, 2, f'개별공시지가 조회결과 - {p["location_full"]}',
                 Font(name='맑은 고딕', size=13, bold=True, color='FFFFFF'),
                 fill_title, Alignment(horizontal='center', vertical='center'))
        for c in range(2, 8):
            ws.cell(row=1, column=c).fill = fill_title
        ws.row_dimensions[1].height = 35

        # 정보 박스
        r = 3
        set_cell(ws, r, 2, '소재지', font_label, fill_header, align_left, border_all)
        ws.merge_cells(f'C{r}:G{r}')
        set_cell(ws, r, 3, p["location_full"], font_data, None, align_left, border_all)
        for cc in range(4, 8): ws.cell(row=r, column=cc).border = border_all
        r += 1

        set_cell(ws, r, 2, f'{data.get("audit_date", "2025")[:4]}년 공시지가', font_label, fill_header, align_left, border_all)
        ws.merge_cells(f'C{r}:G{r}')
        set_cell(ws, r, 3, p["price_2025"],
                 Font(name='맑은 고딕', size=11, bold=True, color='CC0000'),
                 fill_yellow, align_left, border_all)
        for cc in range(4, 8):
            ws.cell(row=r, column=cc).fill = fill_yellow
            ws.cell(row=r, column=cc).border = border_all
        r += 1

        set_cell(ws, r, 2, '총 조회건수', font_label, fill_header, align_left, border_all)
        ws.merge_cells(f'C{r}:G{r}')
        set_cell(ws, r, 3, f'{p["total_records"]}개', font_data, None, align_left, border_all)
        for cc in range(4, 8): ws.cell(row=r, column=cc).border = border_all

        # 데이터 테이블
        r += 2
        set_cell(ws, r, 2, '연도별 공시지가 추이 (최근 6개년)',
                 Font(name='맑은 고딕', size=11, bold=True, color='1A5276'))
        r += 1

        hdrs = ['가격기준년도', '토지소재지', '지번', '개별공시지가', '기준일자', '공시일자']
        for ci, h in enumerate(hdrs):
            set_cell(ws, r, 2+ci, h, font_header, fill_header, align_center, border_all)
        r += 1

        for ri, row_data in enumerate(p["data"][:6]):
            bg = fill_yellow if ri == 0 else (fill_light if ri % 2 == 0 else None)
            for ci, val in enumerate(row_data):
                al = align_right if ci == 3 else align_center
                nf = '#,##0' if ci == 3 else None
                set_cell(ws, r, 2+ci, val, font_data, bg, al, border_all, nf)
            r += 1

        # 이미지 삽입
        r += 2
        set_cell(ws, r, 2, '조회 결과 캡처 이미지',
                 Font(name='맑은 고딕', size=11, bold=True, color='1A5276'))
        r += 1

        img_path = os.path.join(img_dir, p["img_file"])
        if os.path.exists(img_path):
            img = XlImage(img_path)
            img.width = 750
            img.height = 340
            ws.add_image(img, f'B{r}')
            for _ in range(18):
                ws.row_dimensions[r].height = 20
                r += 1

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    data = json.load(sys.stdin)
    result = create_workpaper(data)
    print(f"✅ 워크페이퍼 생성: {result}")
    print(f"   파일 크기: {os.path.getsize(result):,} bytes")
